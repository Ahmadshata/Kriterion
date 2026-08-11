"""CSV, Markdown, HTML, and Excel report generation."""

from __future__ import annotations

import csv
import json
import math
import re
import shutil
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from kriterion import config
from kriterion.dates import format_date, months_to_years
from kriterion.experience import Role
from kriterion.scoring import (
    build_dynamic_headers,
    build_verdict_reasons,
    excel_result_label,
    row_for_result,
)
from kriterion.synonyms import normalize_tool_name


_COPILOT_ICON_SOURCE = (
    Path(__file__).resolve().parent.parent / "assets" / "GitHub-Copilot-Blink.gif"
)
_COPILOT_ICON_FILENAME = _COPILOT_ICON_SOURCE.name


_TOOL_ICON_SOURCE = Path(__file__).resolve().parent.parent / "assets" / "tools"
_TOOL_ICON_ALIASES = {
    "elastic": "elk",
}
_TOOL_ICON_ENLARGED = {"docker", "zabbix"}
_TOOL_LABELS = {
    "argocd": "Argo CD",
    "aws": "AWS",
    "azure devops": "Azure DevOps",
    "gcp": "GCP",
    "gitlab": "GitLab",
    "github actions": "GitHub Actions",
    "kubernetes": "Kubernetes",
    "opentelemetry": "OpenTelemetry",
}


def _tool_slug(tool_name: str) -> str:
    canonical = normalize_tool_name(tool_name)
    return re.sub(r"[^a-z0-9]+", "_", canonical).strip("_")


def _tool_icon_slug(tool_name: str) -> str:
    """Resolve icons, preferring concatenated multi-word filenames."""
    canonical = normalize_tool_name(tool_name)
    compact = re.sub(r"[^a-z0-9]+", "", canonical)
    candidates = [
        _TOOL_ICON_ALIASES.get(canonical, ""),
        compact,
        _tool_slug(canonical),
    ]
    icon_files = {
        path.stem.lower(): path.stem.lower() for path in _TOOL_ICON_SOURCE.glob("*.png")
    }
    for candidate in candidates:
        if candidate and candidate.lower() in icon_files:
            return icon_files[candidate.lower()]
    return ""


def _tool_icon_images_html(tool_name: str) -> str:
    """Render theme-aware icon images when a -white dark-mode asset exists."""
    slug = _tool_icon_slug(tool_name)
    if not slug:
        return ""
    enlarged_class = (
        f" tool-icon-enlarged tool-icon-{slug}" if slug in _TOOL_ICON_ENLARGED else ""
    )
    icon_files = {path.stem.lower() for path in _TOOL_ICON_SOURCE.glob("*.png")}
    white_slug = f"{slug}-white"
    if white_slug in icon_files:
        return (
            f'<img class="tool-icon-light{enlarged_class}" src="tools/{_html_escape(slug)}.png" alt="">'
            f'<img class="tool-icon-dark{enlarged_class}" src="tools/{_html_escape(white_slug)}.png" alt="">'
        )
    class_attr = f' class="{enlarged_class.strip()}"' if enlarged_class else ""
    return f'<img{class_attr} src="tools/{_html_escape(slug)}.png" alt="">'


def _tool_icon_html(tool_name: str) -> str:
    """Render a configured tool icon, with a readable fallback when unavailable."""
    images = _tool_icon_images_html(tool_name)
    if images:
        return f'<span class="kw-tool-icon" aria-hidden="true">{images}</span>'
    fallback = "".join(part[0] for part in tool_name.split() if part)[:2].upper() or "?"
    return (
        '<span class="kw-tool-icon kw-tool-icon-fallback" aria-hidden="true">'
        f"{_html_escape(fallback)}</span>"
    )


def _tool_label(tool_name: str) -> str:
    canonical = normalize_tool_name(tool_name)
    return _TOOL_LABELS.get(canonical, canonical.title())


def _tool_filter_icon_html(tool_name: str) -> str:
    return _tool_icon_images_html(tool_name)


def _role_label(role: Role) -> str:
    """Return the career-history label shown in text and HTML reports."""
    if role.company:
        return f"{role.title} @ {role.company}"
    return role.title


def _role_tenure_years(role: Role) -> float:
    """Return calendar tenure for display, independent of overlap allocation."""
    month_span = (role.end.year - role.start.year) * 12 + (
        role.end.month - role.start.month
    )
    return months_to_years(max(1, month_span))


def _copy_tool_icons(output_dir: Path) -> None:
    """Copy report tool icons using lowercase, URL-safe filenames."""
    if _TOOL_ICON_SOURCE.is_dir():
        destination = output_dir / "tools"
        destination.mkdir(parents=True, exist_ok=True)
        for icon_path in _TOOL_ICON_SOURCE.glob("*.png"):
            shutil.copy2(icon_path, destination / icon_path.name.lower())


def write_csv(results: List[Dict[str, object]], output_path: Path) -> None:
    headers = build_dynamic_headers()
    with output_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in results:
            w.writerow(row_for_result(r, headers))


def write_report(results: List[Dict[str, object]], output_path: Path) -> None:
    total = len(results)
    passed = sum(
        1 for r in results if bool(r["passed"]) and not bool(r.get("ambiguity"))
    )
    ambiguous = sum(1 for r in results if bool(r.get("ambiguity")))
    failed = total - passed - ambiguous

    lines: List[str] = []
    lines.append("# CV Screening Report\n")
    lines.append("## Summary")
    lines.append(f"- Total CVs: {total}")
    lines.append(f"- Passed: {passed}")
    lines.append(f"- Failed: {failed}")
    lines.append(f"- Ambiguous: {ambiguous}\n")

    lines.append("## Active Screening Criteria")
    lines.append(
        f"- Required keywords in Experience: {', '.join(sorted(config.REQUIRED_EXPERIENCE_KEYWORDS))}"
    )
    lines.append(f"- Minimum DevOps experience: {config.MIN_DEVOPS_YEARS} years\n")

    for r in results:
        layout_ambiguity = bool(r.get("layout_ambiguity"))
        lines.append(f"## {r['file']}")
        lines.append(f"- Result: {excel_result_label(r)}")
        lines.append(f"- Confidence Score: {r.get('score', 0)}/100")
        if r["used_ocr"]:
            lines.append("- Note: OCR fallback used for text extraction.")
        detected_tools = r.get("detected_tools", [])
        if isinstance(detected_tools, list) and detected_tools:
            lines.append(
                "- Detected tools in Experience: "
                + ", ".join(_tool_label(str(tool)) for tool in detected_tools)
            )

        required_evidence: Dict[str, Optional[Tuple[int, str]]] = r["required_evidence"]  # type: ignore
        evidence_details: Dict[str, Dict[str, object]] = r.get(
            "required_evidence_details", {}
        )  # type: ignore
        lines.append("- Required keywords evidence (Experience):")
        for kw in sorted(config.REQUIRED_EXPERIENCE_KEYWORDS):
            if layout_ambiguity:
                lines.append(
                    f"  - {kw}: Withheld — multi-column reading order requires review"
                )
                continue
            ev = required_evidence.get(kw.lower())
            detail = evidence_details.get(kw.lower(), {})
            if ev:
                relationship = str(detail.get("relationship", "direct")).replace(
                    "_", " "
                )
                citations = _evidence_citations(detail, ev, kw)
                lines.append(
                    f"  - {kw}: Yes ({relationship}; {len(citations)} "
                    f"citation{'s' if len(citations) != 1 else ''})"
                )
                for citation_index, citation in enumerate(citations, start=1):
                    matched_term = str(citation.get("matched_term", kw))
                    lines.append(
                        f"    - Citation {citation_index}: matched {matched_term}; "
                        f"page {citation['page']}"
                    )
                    snippet = str(citation.get("snippet", ""))
                    lines.append(
                        "      Snippet:\n\n      " + snippet.replace("\n", "\n      ")
                    )
            elif detail.get("needs_review"):
                lines.append(
                    f"  - {kw}: Ambiguous related term "
                    f"({detail.get('matched_term')}; manual review required)"
                )
            else:
                lines.append(f"  - {kw}: No")

        if layout_ambiguity:
            lines.append(
                "- DevOps years counted: Withheld — extraction layout requires review"
            )
        else:
            lines.append(
                f"- DevOps years counted (unique, overlap-safe): {r['devops_years']}"
            )
        lines.append(
            f"- DevOps pass (>= {config.MIN_DEVOPS_YEARS} years AND no ambiguity): {'Yes' if r['devops_pass'] else 'No'}"
        )
        lines.append(f"- Date ambiguity: {'Yes' if r.get('date_ambiguity') else 'No'}")
        lines.append(
            f"- Semantic ambiguity: {'Yes' if r.get('semantic_ambiguity') else 'No'}"
        )
        lines.append(
            f"- Extraction-layout ambiguity: {'Yes' if r.get('layout_ambiguity') else 'No'}"
        )

        roles: List[Role] = r["devops_roles"]  # type: ignore
        if layout_ambiguity:
            lines.append(
                "- DevOps roles counted: Withheld — extraction layout requires review"
            )
        elif roles:
            lines.append("- DevOps roles counted:")
            for role in roles:
                role_years = _role_tenure_years(role)
                role_period = f"{format_date(role.start)} to {format_date(role.end)}"
                lines.append(
                    f"  - {_role_label(role)} ({role_period}): {role_years} years"
                )
        else:
            lines.append("- DevOps roles counted: None")

        excl: List[str] = r["education_program_entries"]  # type: ignore
        if excl:
            lines.append("- Education program entries (not counted as experience):")
            for e in excl:
                lines.append(f"  - {e}")
        else:
            lines.append("- Education program entries: None")

        freelance_excluded: List[str] = r.get("freelance_entries_excluded", [])  # type: ignore
        if freelance_excluded:
            lines.append("- Freelance entries excluded by profile:")
            for entry in freelance_excluded:
                lines.append(f"  - {entry}")

        if r.get("excluded_company"):
            lines.append(f"- EXCLUDED COMPANY: {r['excluded_company']}")
        if r.get("excluded_university"):
            lines.append(f"- EXCLUDED UNIVERSITY: {r['excluded_university']}")
        if config.PREFERRED_PROGRAM_PATTERNS:
            if r.get("preferred_program"):
                lines.append(f"- Preferred program found: {r['preferred_program']}")
            else:
                lines.append("- Preferred program found: No (REQUIRED)")

        lines.append("")

    output_path.write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def _html_escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def _highlight_term(escaped_snippet: str, matched_term: str) -> str:
    escaped_term = _html_escape(matched_term)
    pattern = re.compile(re.escape(escaped_term), re.IGNORECASE)
    return pattern.sub(
        lambda m: f'<mark class="kw-hl">{m.group(0)}</mark>',
        escaped_snippet,
    )


def _evidence_citations(
    detail: Dict[str, object],
    primary: Optional[Tuple[int, str]],
    default_term: str,
) -> List[Dict[str, object]]:
    """Return all citations, falling back to legacy single-citation results."""
    citations = detail.get("citations")
    if isinstance(citations, list):
        valid = [
            citation
            for citation in citations
            if isinstance(citation, dict)
            and citation.get("page") is not None
            and str(citation.get("snippet", "")).strip()
        ]
        if valid:
            return valid
    if primary:
        return [
            {
                "page": primary[0],
                "snippet": primary[1],
                "matched_term": detail.get("matched_term", default_term),
            }
        ]
    return []


def write_html_report(
    results: List[Dict[str, object]],
    output_path: Path,
    cv_folder: Optional[Path] = None,
    *,
    auto_ai_review: bool = True,
    profile: Optional[Dict[str, object]] = None,
    ai_provider: str = "codex",
) -> None:
    ai_provider = str(ai_provider).strip().lower()
    if ai_provider not in {"codex", "copilot"}:
        raise ValueError("ai_provider must be 'codex' or 'copilot'")
    show_ai_usage = ai_provider == "copilot"
    ai_provider_name = "GitHub Copilot" if show_ai_usage else "Codex"
    ai_section_icon = (
        f'<div class="section-icon section-icon-ai"><img class="copilot-icon" src="{_COPILOT_ICON_FILENAME}" alt=""></div>'
        if show_ai_usage
        else '<div class="section-icon section-icon-codex" aria-hidden="true">CX</div>'
    )
    _copy_tool_icons(output_path.parent)
    if _COPILOT_ICON_SOURCE.is_file():
        shutil.copy2(_COPILOT_ICON_SOURCE, output_path.parent / _COPILOT_ICON_FILENAME)
    total = len(results)
    passed = sum(
        1 for r in results if bool(r["passed"]) and not bool(r.get("ambiguity"))
    )
    ambiguous = sum(1 for r in results if bool(r.get("ambiguity")))
    failed = total - passed - ambiguous
    kw_coverage = {}
    for kw in sorted(config.REQUIRED_EXPERIENCE_KEYWORDS):
        found = sum(
            1 for r in results if r["required_evidence"].get(kw.lower()) is not None
        )
        kw_coverage[kw] = {
            "found": found,
            "missing": total - found,
            "pct": round(found / max(total, 1) * 100),
        }

    tool_counts: Dict[str, int] = {}
    for result in results:
        detected_tools = result.get("detected_tools", [])
        if not isinstance(detected_tools, list):
            continue
        candidate_tools = {
            normalize_tool_name(str(value))
            for value in detected_tools
            if str(value).strip()
        }
        for tool in candidate_tools:
            tool_counts[tool] = tool_counts.get(tool, 0) + 1
    required_tool_names = {
        normalize_tool_name(str(tool)) for tool in config.REQUIRED_EXPERIENCE_KEYWORDS
    }
    sorted_tool_counts = sorted(
        tool_counts.items(),
        key=lambda item: (
            0 if item[0].lower() in required_tool_names else 1,
            -item[1],
            _tool_label(item[0]).lower(),
        ),
    )
    tool_filter_chips = "".join(
        f'<button type="button" class="tool-filter-chip" '
        f'data-tool="{_html_escape(_tool_slug(tool))}" aria-pressed="false" '
        f'aria-label="Filter by {_html_escape(_tool_label(tool))}">'
        f"{_tool_filter_icon_html(tool)}<span>{_html_escape(_tool_label(tool))}</span>"
        f"<strong>{count}</strong></button>"
        for tool, count in sorted_tool_counts
    )
    tool_filter_panel = (
        f"""<section class="tool-filter-panel" aria-labelledby="toolFilterTitle">
            <div class="tool-filter-head">
                <div><span class="tool-filter-kicker">Experience technology index</span><h2 id="toolFilterTitle">Filter candidates by tools</h2><p>Select multiple tools to require every selected tool.</p></div>
                <button type="button" class="tool-filter-clear" id="toolFilterClear" disabled>Clear tools</button>
            </div>
            <div class="tool-filter-chips" aria-label="Tool filters">{tool_filter_chips}</div>
            <div class="tool-filter-summary" id="toolFilterSummary" aria-live="polite">Showing all {total} candidates</div>
        </section>"""
        if tool_filter_chips
        else ""
    )

    reason_counts: Dict[str, int] = {}
    for r in results:
        if r["passed"] and not r.get("ambiguity"):
            continue
        for kw, ev in r["required_evidence"].items():
            if ev is None:
                key = f"Missing: {kw}"
                reason_counts[key] = reason_counts.get(key, 0) + 1
        if r["devops_years"] < config.MIN_DEVOPS_YEARS and not r.get("ambiguity"):
            reason_counts["Insufficient experience years"] = (
                reason_counts.get("Insufficient experience years", 0) + 1
            )
        if r.get("date_ambiguity"):
            reason_counts["Date ambiguity"] = reason_counts.get("Date ambiguity", 0) + 1
        if r.get("semantic_ambiguity"):
            reason_counts["Semantic ambiguity"] = (
                reason_counts.get("Semantic ambiguity", 0) + 1
            )
        if r.get("layout_ambiguity"):
            reason_counts["Extraction-layout ambiguity"] = (
                reason_counts.get("Extraction-layout ambiguity", 0) + 1
            )
        if r.get("excluded_company"):
            reason_counts["Excluded company"] = (
                reason_counts.get("Excluded company", 0) + 1
            )
        if r.get("excluded_university"):
            reason_counts["Excluded university"] = (
                reason_counts.get("Excluded university", 0) + 1
            )
    top_reasons = sorted(reason_counts.items(), key=lambda x: -x[1])[:8]
    max_reason_count = top_reasons[0][1] if top_reasons else 1

    cv_base = cv_folder.resolve() if cv_folder else None

    def _cv_url(filename: str) -> str:
        if not cv_base:
            return ""
        from urllib.parse import quote

        return "/cvs/" + quote(filename, safe="")

    def _file_link(filename: str) -> str:
        escaped = _html_escape(filename)
        file_url = _cv_url(filename)
        if file_url:
            return f'<a href="{file_url}" target="_blank" class="file-link" title="Open CV">{escaped}</a>'
        return escaped

    def _citation_role_context(result: Dict[str, object], snippet: str) -> str:
        snippet_lower = snippet.lower()
        result_roles = result.get("devops_roles", [])
        if not isinstance(result_roles, list):
            return ""
        for role in result_roles:
            if not isinstance(role, Role):
                continue
            title_match = bool(role.title and role.title.lower() in snippet_lower)
            company_match = bool(role.company and role.company.lower() in snippet_lower)
            if title_match or company_match:
                return (
                    f"{_role_label(role)} · {format_date(role.start)} — "
                    f"{format_date(role.end)}"
                )
        return ""

    kw_bars_html = ""
    for kw, data in kw_coverage.items():
        pct = data["pct"]
        color = (
            "var(--green)"
            if pct >= 70
            else "var(--amber)"
            if pct >= 40
            else "var(--red)"
        )
        kw_bars_html += f'<div class="chart-row"><span class="chart-label">{_html_escape(kw)}</span><div class="chart-bar-track"><div class="chart-bar-fill" style="width:{pct}%;background:{color}"></div></div><span class="chart-value">{pct}%</span></div>'

    reasons_bars_html = ""
    for reason, count in top_reasons:
        pct = round(count / max(max_reason_count, 1) * 100)
        reasons_bars_html += f'<div class="chart-row"><span class="chart-label">{_html_escape(reason)}</span><div class="chart-bar-track"><div class="chart-bar-fill reason-fill" style="width:{pct}%"></div></div><span class="chart-value">{count}</span></div>'

    table_rows = ""
    evidence_xray_items: List[Dict[str, object]] = []
    for idx, r in enumerate(results):
        result_label = excel_result_label(r)
        score = int(r.get("score", 0))
        status_class = result_label.lower()
        kw_found = sum(1 for v in r["required_evidence"].values() if v is not None)  # type: ignore
        kw_total = len(config.REQUIRED_EXPERIENCE_KEYWORDS)
        reasons = build_verdict_reasons(r)
        layout_ambiguity = bool(r.get("layout_ambiguity"))
        table_years = (
            "Review"
            if layout_ambiguity
            else f"{math.ceil(r['devops_years'] * 2) / 2} yr"
        )
        table_requirements = "Review" if layout_ambiguity else f"{kw_found}/{kw_total}"

        detected_tools = [] if layout_ambiguity else r.get("detected_tools", [])
        normalized_tools = (
            {
                normalize_tool_name(str(tool))
                for tool in detected_tools
                if str(tool).strip()
            }
            if isinstance(detected_tools, list)
            else set()
        )
        tool_slugs = " ".join(_tool_slug(tool) for tool in sorted(normalized_tools))
        candidate_filename = str(r["file"])
        candidate_cv_url = _cv_url(candidate_filename)
        candidate_xray_url = (
            candidate_cv_url.replace("/cvs/", "/xray/", 1)
            if candidate_cv_url and candidate_filename.lower().endswith(".pdf")
            else ""
        )
        roles: List[Role] = r["devops_roles"]  # type: ignore
        table_rows += f"""<tr class="data-row row-{status_class}" data-status="{status_class}" data-tools="{_html_escape(tool_slugs)}" data-idx="{idx}" role="button" tabindex="0" aria-expanded="false" aria-label="Review candidate {_html_escape(str(r["file"]))}">
            <td class="cell-name">{_file_link(str(r["file"]))}</td>
            <td><span class="pill pill-{status_class}">{result_label}</span></td>
            <td><div class="score-bar-cell"><div class="score-bar-track"><div class="score-bar-fill score-fill-{status_class}" style="width:{score}%"></div></div><span class="score-num">{score}</span></div></td>
            <td>{table_years}</td>
            <td>{table_requirements}</td>
        </tr>"""

        kw_items = ""
        required_evidence: Dict[str, Optional[Tuple[int, str]]] = r["required_evidence"]  # type: ignore
        evidence_details: Dict[str, Dict[str, object]] = r.get(
            "required_evidence_details", {}
        )  # type: ignore
        for kw in sorted(config.REQUIRED_EXPERIENCE_KEYWORDS):
            ev = required_evidence.get(kw.lower())
            detail = evidence_details.get(kw.lower(), {})
            tool_heading = (
                '<div class="kw-title">'
                f"{_tool_icon_html(kw)}"
                '<div class="kw-title-copy">'
                f'<span class="kw-name">{_html_escape(kw)}</span>'
                '<span class="kw-kicker">Required experience</span>'
                "</div></div>"
            )
            if layout_ambiguity:
                kw_items += f"""<div class="kw-item kw-review">
                    <div class="kw-header">{tool_heading}<span class="kw-badge review">AI review</span></div>
                    <p class="empty-state">Deterministic citations withheld because the multi-column reading order is unreliable.</p>
                </div>"""
                continue
            if ev:
                relationship = str(detail.get("relationship", "direct"))
                citations = _evidence_citations(detail, ev, kw)
                citation_count = len(citations)
                evidence_label = (
                    f"{relationship.replace('_', ' ')} · {citation_count} "
                    f"citation{'s' if citation_count != 1 else ''}"
                )
                citation_items = ""
                for citation_index, citation in enumerate(citations, start=1):
                    matched_term = str(citation.get("matched_term", kw))
                    citation_snippet = str(citation.get("snippet", ""))
                    snippet = _html_escape(citation_snippet)
                    snippet_hl = _highlight_term(snippet, matched_term)
                    xray_id = f"xray-{idx}-{_tool_slug(kw)}-{citation_index}"
                    evidence_xray_items.append(
                        {
                            "id": xray_id,
                            "candidate": candidate_filename,
                            "requirement": _tool_label(kw),
                            "relationship": relationship.replace("_", " "),
                            "page": int(citation["page"]),
                            "matchedTerm": matched_term,
                            "snippet": citation_snippet,
                            "roleContext": _citation_role_context(r, citation_snippet),
                            "sourceUrl": candidate_cv_url,
                            "previewUrl": candidate_xray_url,
                            "previewAvailable": bool(candidate_xray_url),
                        }
                    )
                    citation_items += f"""<article class="kw-citation">
                        <div class="kw-citation-head"><span class="kw-citation-index">Evidence {citation_index:02d}</span><div class="kw-citation-source"><span class="kw-relation">Page {_html_escape(str(citation["page"]))} · Matched: {_html_escape(matched_term)}</span><button type="button" class="xray-open-btn" data-xray-id="{xray_id}"><span aria-hidden="true">⌖</span> Inspect source</button></div></div>
                        <pre class="kw-snippet">{snippet_hl}</pre>
                    </article>"""
                kw_items += f"""<div class="kw-item kw-found">
                    <div class="kw-header">{tool_heading}<span class="kw-badge found">{_html_escape(evidence_label)}</span></div>
                    <div class="kw-citations" aria-label="{citation_count} evidence citations">{citation_items}</div>
                </div>"""
            elif detail.get("needs_review"):
                review_matched_term = str(detail.get("matched_term", ""))
                matched_term = _html_escape(review_matched_term)
                review_snippet = str(detail.get("snippet", ""))
                snippet = _html_escape(review_snippet)
                snippet_hl = _highlight_term(snippet, review_matched_term)
                xray_id = f"xray-{idx}-{_tool_slug(kw)}-review"
                xray_page = int(detail.get("page") or 1)
                evidence_xray_items.append(
                    {
                        "id": xray_id,
                        "candidate": candidate_filename,
                        "requirement": _tool_label(kw),
                        "relationship": str(
                            detail.get("relationship", "related evidence")
                        ).replace("_", " "),
                        "page": xray_page,
                        "matchedTerm": review_matched_term,
                        "snippet": review_snippet,
                        "roleContext": _citation_role_context(r, review_snippet),
                        "sourceUrl": candidate_cv_url,
                        "previewUrl": candidate_xray_url,
                        "previewAvailable": bool(candidate_xray_url),
                    }
                )
                kw_items += f"""<div class="kw-item kw-review">
                    <div class="kw-header">{tool_heading}<span class="kw-badge review">review needed</span></div>
                    <article class="kw-citation kw-citation-review"><div class="kw-citation-head"><span class="kw-citation-index">Related evidence</span><div class="kw-citation-source"><span class="kw-relation">Page {xray_page} · Matched: {matched_term}</span><button type="button" class="xray-open-btn" data-xray-id="{xray_id}"><span aria-hidden="true">⌖</span> Inspect source</button></div></div><pre class="kw-snippet">{snippet_hl}</pre></article>
                </div>"""
            else:
                kw_items += f"""<div class="kw-item kw-missing">
                    <div class="kw-header">{tool_heading}<span class="kw-badge missing">missing</span></div>
                </div>"""

        exp_items = ""
        if layout_ambiguity:
            exp_items = '<p class="empty-state">Career history withheld because the multi-column reading order is unreliable. Use the AI verdict and verify against the original CV.</p>'
        elif roles:
            newest_roles = sorted(
                roles,
                key=lambda role: (role.start, role.end),
                reverse=True,
            )
            for role_index, role in enumerate(newest_roles):
                ry = _role_tenure_years(role)
                has_next_role = role_index < len(newest_roles) - 1
                connected_class = " exp-item-connected" if has_next_role else ""
                connector = (
                    '<div class="exp-connector" aria-hidden="true">'
                    '<span class="exp-particle"></span>'
                    '<span class="exp-particle"></span>'
                    '<span class="exp-particle"></span>'
                    "</div>"
                    if has_next_role
                    else ""
                )
                exp_items += f"""<div class="exp-item{connected_class}">
                    <div class="exp-dot"></div>
                    <div class="exp-content">
                        <div class="exp-title">{_html_escape(_role_label(role))}</div>
                        <div class="exp-meta">{format_date(role.start)} — {format_date(role.end)} &middot; {ry} yr tenure</div>
                    </div>
                    {connector}
                </div>"""
        else:
            exp_items = '<p class="empty-state">No DevOps roles detected</p>'

        flags_items = ""
        edu_entries: List[str] = r.get("education_program_entries", [])  # type: ignore
        if edu_entries:
            for e in edu_entries:
                flags_items += f'<div class="flag-item flag-info"><span class="flag-icon">\U0001f393</span><span>Listed as experience but classified as education: {_html_escape(e)}</span></div>'
        freelance_entries: List[str] = r.get("freelance_entries_excluded", [])  # type: ignore
        for entry in freelance_entries:
            flags_items += f'<div class="flag-item flag-info"><span class="flag-icon">◌</span><span>Freelance experience excluded by profile: {_html_escape(entry)}</span></div>'
        if r.get("layout_ambiguity"):
            flags_items += '<div class="flag-item flag-info"><span class="flag-icon">⚠</span><span>Multi-column extraction produced an unreliable experience reading order; AI-assisted review and a human decision are required.</span></div>'
        if r.get("excluded_company"):
            flags_items += f'<div class="flag-item flag-danger"><span class="flag-icon">\U0001f6ab</span><span>Excluded company: {_html_escape(str(r["excluded_company"]))}</span></div>'
        if r.get("excluded_university"):
            flags_items += f'<div class="flag-item flag-danger"><span class="flag-icon">\U0001f6ab</span><span>Excluded university: {_html_escape(str(r["excluded_university"]))}</span></div>'
        if config.PREFERRED_PROGRAM_PATTERNS:
            if r.get("preferred_program"):
                flags_items += f'<div class="flag-item flag-success"><span class="flag-icon">✓</span><span>Preferred program: {_html_escape(str(r["preferred_program"]))}</span></div>'
            else:
                flags_items += '<div class="flag-item flag-danger"><span class="flag-icon">✗</span><span>No preferred program found (required)</span></div>'
        if not flags_items:
            flags_items = '<p class="empty-state">No flags</p>'

        verdict_icon = (
            "✓"
            if result_label == "PASS"
            else ("⚠" if result_label == "AMBIGUOUS" else "✗")
        )
        reason_items = "".join(
            f'<li class="verdict-item">{_html_escape(reason)}</li>'
            for reason in reasons
        )

        ai_review_section = ""
        if bool(r.get("ambiguity")):
            ai_review_section = f"""<section class="review-section ai-review-section">
                <div class="review-section-head">
                    {ai_section_icon}
                    <div>
                        <h4>AI verdict</h4>
                        <p>A second opinion from {ai_provider_name} for the unresolved evidence</p>
                    </div>
                </div>
                <div class="semantic-review-container" data-file="{_html_escape(str(r["file"]))}" data-idx="{idx}">
                    <div class="ai-scope"><span>{ai_provider_name}</span><span>Ambiguity only</span><span>Work experience only</span><span>Human decides</span></div>
                    <button class="semantic-review-btn" onclick="requestAmbiguityVerdict(this)">Generate AI verdict</button>
                    <div class="semantic-review-output"></div>
                </div>
            </section>"""

        interview_architect_section = ""
        if result_label == "PASS":
            interview_architect_section = f"""<section class="review-section interview-architect-section">
            <div class="review-section-head interview-architect-head">
                {ai_section_icon}
                <div>
                    <h4>Interview Architect</h4>
                    <p>{ai_provider_name} identifies each candidate issue and creates one evidence-led question for it</p>
                </div>
            </div>
            <div class="interview-architect-container" data-file="{_html_escape(str(r["file"]))}" data-idx="{idx}">
                <div class="interview-scope"><span>{ai_provider_name}</span><span>Ambiguous evidence</span><span>Strong claims</span><span>Timeline overlaps</span><span>Career gaps</span></div>
                <button type="button" class="interview-architect-btn" onclick="requestInterviewPlan(this)">Analyze issues &amp; build questions</button>
                <div class="interview-architect-output" aria-live="polite"></div>
            </div>
        </section>"""

        outcome_copy = (
            "Needs a reviewer to resolve uncertain evidence"
            if result_label == "AMBIGUOUS"
            else (
                "Meets the configured screening requirements"
                if result_label == "PASS"
                else "Does not meet one or more screening requirements"
            )
        )
        years_display = (
            "Review" if layout_ambiguity else str(math.ceil(r["devops_years"] * 2) / 2)
        )
        requirements_display = (
            "Review" if layout_ambiguity else f"{kw_found}/{kw_total}"
        )
        evidence_summary = (
            "Deterministic citations withheld pending AI and human review"
            if layout_ambiguity
            else f"{kw_found} of {kw_total} required technologies confirmed in work experience"
        )
        career_summary = (
            "Extracted timeline requires AI and human review"
            if layout_ambiguity
            else f"{len(roles)} qualifying role{'s' if len(roles) != 1 else ''} used in the calculation"
        )

        table_rows += f"""<tr class="detail-row" data-idx="{idx}" data-status="{status_class}" style="display:none">
            <td colspan="5">
                <div class="candidate-review">
                    <header class="review-hero review-hero-{status_class}">
                        <div class="review-outcome">
                            <div class="outcome-icon">{verdict_icon}</div>
                            <div>
                                <div class="review-eyebrow">Screening outcome</div>
                                <h3>{result_label}</h3>
                                <p>{outcome_copy}</p>
                            </div>
                        </div>
                        <div class="review-metrics">
                            <div class="review-metric"><strong>{score}</strong><span>Score</span></div>
                            <div class="review-metric"><strong>{years_display}</strong><span>Years</span></div>
                            <div class="review-metric"><strong>{requirements_display}</strong><span>Requirements</span></div>
                        </div>
                    </header>
                    <div class="review-body">
                        <div class="lab-application-note" hidden>Kriterion Lab changed the displayed outcome using this report’s verified evidence. The detailed evidence and profile YAML remain unchanged.</div>
                        <section class="decision-rationale rationale-{status_class}">
                            <div class="rationale-heading">
                                <span>Why this decision</span>
                                <small>Deterministic screening evidence</small>
                            </div>
                            <ul>{reason_items}</ul>
                        </section>
                        <section class="review-section evidence-section">
                            <div class="review-section-head">
                                <div class="section-icon">01</div>
                                <div>
                                    <h4>Requirement evidence</h4>
                                    <p>{evidence_summary}</p>
                                </div>
                            </div>
                            <div class="keyword-grid">{kw_items}</div>
                        </section>
                        <div class="review-split">
                            <section class="review-section">
                                <div class="review-section-head">
                                    <div class="section-icon">02</div>
                                    <div>
                                        <h4>Relevant career history</h4>
                                        <p>{career_summary}</p>
                                    </div>
                                </div>
                                <div class="timeline">{exp_items}</div>
                            </section>
                            <section class="review-section">
                                <div class="review-section-head">
                                    <div class="section-icon">03</div>
                                    <div>
                                        <h4>Screening notes</h4>
                                        <p>Education, exclusions, and preferred-program checks</p>
                                    </div>
                                </div>
                                <div class="screening-notes">{flags_items}</div>
                            </section>
                        </div>
                        {ai_review_section}
                        {interview_architect_section}
                    </div>
                </div>
            </td>
        </tr>"""

    pass_pct = passed / max(total, 1) * 100
    fail_pct = failed / max(total, 1) * 100
    ambiguous_pct = ambiguous / max(total, 1) * 100
    pass_rate = round(pass_pct)

    def _profile_list(key: str) -> List[str]:
        values = (profile or {}).get(key, [])
        if not isinstance(values, list):
            return []
        return [str(value) for value in values if str(value).strip()]

    def _profile_short_names(key: str) -> List[str]:
        short_names = [value for value in _profile_list(key) if len(value.split()) == 1]
        return [
            value.upper() if value.isalpha() and len(value) <= 5 else value.title()
            for value in short_names
        ]

    def _restriction_group(label: str, values: List[str], chip_class: str) -> str:
        if not values:
            return ""
        chips = "".join(
            f'<span class="chip {chip_class}">{_html_escape(value)}</span>'
            for value in values
        )
        return (
            f'<div class="criteria-section"><h4>{_html_escape(label)}</h4>'
            f'<div class="chips">{chips}</div></div>'
        )

    min_years = f"{config.MIN_DEVOPS_YEARS:g}+ years"
    restriction_groups_html = "".join(
        [
            _restriction_group(
                "Required in Experience",
                sorted(config.REQUIRED_EXPERIENCE_KEYWORDS),
                "chip-required",
            ),
            _restriction_group("Minimum Experience", [min_years], "chip-rule"),
            _restriction_group(
                "Freelance Experience",
                [
                    "Included"
                    if bool((profile or {}).get("include_freelance_experience", True))
                    else "Excluded"
                ],
                "chip-rule",
            ),
            _restriction_group(
                "Minimum Score",
                [f"{config.MIN_SCORE}/100"] if config.MIN_SCORE is not None else [],
                "chip-rule",
            ),
            _restriction_group(
                "One Preferred Program Required",
                _profile_list("preferred_programs"),
                "chip-preferred",
            ),
            _restriction_group(
                "Excluded Companies",
                _profile_list("excluded_companies"),
                "chip-excluded",
            ),
            _restriction_group(
                "Excluded Universities",
                _profile_list("excluded_universities"),
                "chip-excluded",
            ),
            _restriction_group(
                "Not Counted as Experience",
                _profile_short_names("education_programs"),
                "chip-muted",
            ),
        ]
    )

    criterion_lab_rules = sorted(config.REQUIRED_EXPERIENCE_KEYWORDS)
    criterion_lab_candidates = []
    preferred_program_required = bool(_profile_list("preferred_programs"))
    for result in results:
        required_evidence = result.get("required_evidence", {})
        evidence_details = result.get("required_evidence_details", {})
        lab_evidence: Dict[str, str] = {}
        for keyword in criterion_lab_rules:
            if bool(result.get("layout_ambiguity")):
                state = "review"
            elif required_evidence.get(keyword.lower()) is not None:  # type: ignore
                state = "found"
            elif evidence_details.get(keyword.lower(), {}).get("needs_review"):  # type: ignore
                state = "review"
            else:
                state = "missing"
            lab_evidence[keyword] = state

        fixed_failures = []
        if result.get("excluded_company"):
            fixed_failures.append("Excluded company")
        if result.get("excluded_university"):
            fixed_failures.append("Excluded university")
        if preferred_program_required and not result.get("preferred_program"):
            fixed_failures.append("Preferred program missing")
        if (
            config.MIN_SCORE is not None
            and int(result.get("score", 0)) < config.MIN_SCORE
        ):
            fixed_failures.append("Score threshold")

        criterion_lab_candidates.append(
            {
                "name": str(result.get("file", "Candidate")),
                "years": float(result.get("devops_years", 0)),
                "dateReview": bool(result.get("date_ambiguity")),
                "layoutReview": bool(result.get("layout_ambiguity")),
                "evidence": lab_evidence,
                "fixedFailures": fixed_failures,
            }
        )

    criterion_lab_max_years = max(
        5,
        math.ceil(config.MIN_DEVOPS_YEARS + 2),
        math.ceil(
            max(
                (float(result.get("devops_years", 0)) for result in results),
                default=0,
            )
            + 1
        ),
    )
    criterion_lab_data_json = json.dumps(
        {
            "minimumYears": config.MIN_DEVOPS_YEARS,
            "rules": criterion_lab_rules,
            "candidates": criterion_lab_candidates,
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )
    criterion_lab_data_json = (
        criterion_lab_data_json.replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
    )
    evidence_xray_data_json = json.dumps(
        evidence_xray_items,
        ensure_ascii=False,
        separators=(",", ":"),
    )
    evidence_xray_data_json = (
        evidence_xray_data_json.replace("&", "\\u0026")
        .replace("<", "\\u003c")
        .replace(">", "\\u003e")
    )

    criterion_lab_rule_controls = "".join(
        f"""<div class="criterion-rule" data-criterion-rule="{_html_escape(keyword)}">
            <div class="criterion-rule-name">{_tool_icon_html(keyword)}<span>{_html_escape(_tool_label(keyword))}</span></div>
            <div class="criterion-mode-group" role="group" aria-label="Mode for {_html_escape(_tool_label(keyword))}">
                <button type="button" class="criterion-mode-btn is-active" data-mode="required" aria-pressed="true">Required</button>
                <button type="button" class="criterion-mode-btn" data-mode="preferred" aria-pressed="false">Preferred</button>
                <button type="button" class="criterion-mode-btn" data-mode="off" aria-pressed="false">Off</button>
            </div>
        </div>"""
        for keyword in criterion_lab_rules
    )
    criterion_lab_impact_rows = (
        '<div class="criterion-impact-row" data-impact-rule="__years__">'
        '<div><strong>Minimum experience</strong><span class="criterion-impact-copy">Calculating impact…</span></div>'
        '<div class="criterion-impact-track"><i></i></div></div>'
        + "".join(
            f'<div class="criterion-impact-row" data-impact-rule="{_html_escape(keyword)}">'
            f'<div><strong>{_html_escape(_tool_label(keyword))}</strong><span class="criterion-impact-copy">Calculating impact…</span></div>'
            '<div class="criterion-impact-track"><i></i></div></div>'
            for keyword in criterion_lab_rules
        )
    )

    profile_role = _html_escape(str((profile or {}).get("role", "Target role")))
    podium_skills_html = "".join(
        f'<div class="demo-skill"><span>{_html_escape(keyword)}</span>'
        f'<b><i style="--skill:{data["pct"]}%"></i></b>'
        f"<em>{data['pct']}%</em></div>"
        for keyword, data in list(kw_coverage.items())[:3]
    )
    ai_usage_overview = (
        """<div class="ai-usage-overview" id="aiUsageOverview" aria-live="polite">
        <div class="ai-usage-overview-copy">
            <span>Ambiguity-review AI credits</span>
            <strong id="aiUsageTotal">No AI reviews loaded</strong>
        </div>
        <small id="aiUsageBreakdown">Exact Copilot credit usage appears after ambiguous CV verdicts are generated or loaded.</small>
    </div>"""
        if show_ai_usage
        else ""
    )

    html = f"""<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Kriterion — Report</title>
<link rel="icon" type="image/png" sizes="48x48" href="icon.png">
<link rel="apple-touch-icon" href="icon.png">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
:root{{
--bg:#0b0e14;--bg2:#12161f;--bg3:#1a1f2e;--bg4:#222838;
--border:#2a3041;--border2:#353d52;
--text:#e2e8f4;--text2:#8892a8;--text3:#5c6680;
--accent:#7c5cfc;--accent2:#9d7dff;--accent-g:rgba(124,92,252,.12);
--green:#22c55e;--green-g:rgba(34,197,94,.1);
--red:#f43f5e;--red-g:rgba(244,63,94,.1);
--amber:#f59e0b;--amber-g:rgba(245,158,11,.1);
--radius:14px;--radius-sm:8px;
--font:-apple-system,BlinkMacSystemFont,'Inter','Segoe UI',sans-serif;
--mono:'JetBrains Mono','Fira Code','SF Mono',monospace;
}}
[data-theme="light"]{{
--bg:#f4f6fa;--bg2:#fff;--bg3:#fff;--bg4:#f0f2f8;
--border:#e0e4ee;--border2:#d0d5e2;
--text:#1a1f2e;--text2:#5c6680;--text3:#8892a8;
}}
body{{font-family:var(--font);font-size:16px;background:var(--bg);color:var(--text);min-height:100vh}}
a{{color:inherit}}

/* Header — 3-row structured layout */
.header{{background:var(--bg2);border-bottom:1px solid var(--border);padding:1.8rem 2.5rem 1.5rem;display:flex;flex-direction:column;gap:1.4rem}}
.header-top{{display:flex;justify-content:space-between;align-items:center}}
.header-brand{{display:flex;align-items:center;gap:1rem}}
.logo-icon{{width:76px;height:76px;border-radius:14px;overflow:hidden;animation:brand-mark-in .95s cubic-bezier(.2,.8,.2,1) both}}.logo-icon img{{width:100%;height:100%;object-fit:contain}}
.logo-text{{font-size:2rem;font-weight:760;letter-spacing:-.045em;animation:brand-name-in .85s .14s cubic-bezier(.2,.8,.2,1) both}}
.theme-btn{{background:var(--bg3);border:1px solid var(--border);border-radius:9px;padding:.55rem 1rem;cursor:pointer;font-size:.85rem;transition:all .2s;white-space:nowrap}}
.theme-btn:hover{{border-color:var(--accent)}}
@keyframes brand-mark-in{{from{{opacity:0;transform:translateY(9px) scale(.86) rotate(-3deg)}}to{{opacity:1;transform:none}}}}
@keyframes brand-name-in{{from{{opacity:0;transform:translateX(-10px);letter-spacing:.035em}}to{{opacity:1;transform:none;letter-spacing:-.045em}}}}

/* Screening summary: intake, outcomes, and pass-rate ring */
.header-middle{{position:relative;display:grid;grid-template-columns:minmax(205px,.58fr) minmax(470px,1.45fr) minmax(275px,.74fr);align-items:stretch;padding:1.45rem .2rem 1.7rem;border-top:1px solid var(--border);border-bottom:1px solid var(--border)}}
.stat-box{{--status:var(--accent);appearance:none;background:transparent;color:inherit;font:inherit;cursor:pointer;transition:background .22s,box-shadow .22s;border:0}}
.stat-box:hover{{background:color-mix(in srgb,var(--status) 5%,transparent)}}
.stat-box:focus-visible{{outline:2px solid var(--status);outline-offset:-3px}}
.stat-box.stat-active{{background:linear-gradient(180deg,transparent,color-mix(in srgb,var(--status) 8%,transparent));box-shadow:inset 0 -2px var(--status)}}
.screening-intake{{--status:var(--accent);display:grid;align-content:center;min-height:200px;padding:1.1rem clamp(1.2rem,2.2vw,2.4rem) 1.1rem .45rem;text-align:left;border-right:1px solid var(--border)}}
.summary-kicker{{color:var(--accent2);font-size:.7rem;font-weight:750;letter-spacing:.12em;text-transform:uppercase}}
.screening-intake .stat-val{{margin-top:.2rem;font-size:clamp(4rem,6vw,6.4rem);font-weight:780;letter-spacing:-.085em;line-height:.94}}
.screening-intake .stat-lbl{{margin-top:.6rem;color:var(--text2);font-size:.8rem}}
.outcome-grid{{display:grid;grid-template-columns:repeat(3,1fr);align-items:stretch;gap:clamp(.6rem,1.5vw,1.3rem);padding:0 clamp(1.2rem,2.6vw,2.8rem)}}
.outcome-card{{position:relative;display:grid;align-content:center;min-height:200px;padding:1.3rem .8rem 1.05rem 1.45rem;text-align:left;box-shadow:inset 0 1px var(--border),inset 0 -1px var(--border)}}
.outcome-card.stat-pass{{--status:var(--green)}}.outcome-card.stat-fail{{--status:var(--red)}}.outcome-card.stat-amb{{--status:var(--amber)}}
.outcome-bar{{position:absolute;left:0;top:14%;width:3px;height:0;border-radius:1px;background:var(--status);animation:outcome-bar-down 1.65s cubic-bezier(.2,.82,.2,1) forwards}}
.outcome-card:nth-child(2) .outcome-bar{{animation-delay:.12s}}.outcome-card:nth-child(3) .outcome-bar{{animation-delay:.24s}}
.outcome-card .stat-lbl{{color:var(--text2);font-size:.7rem;font-weight:740;letter-spacing:.11em;text-transform:uppercase}}
.outcome-card .stat-val{{margin-top:.42rem;color:var(--status);font-size:clamp(2.2rem,3.4vw,3.5rem);font-weight:750;letter-spacing:-.055em;line-height:1}}
.stat-pct{{margin-top:.55rem;color:var(--text3);font-family:var(--mono);font-size:.7rem}}
.donut-wrap{{display:flex;align-items:center;justify-content:center;gap:.9rem;padding-left:clamp(1.2rem,2.5vw,2.5rem);border-left:1px solid var(--border)}}
.donut{{width:122px;height:122px;position:relative;flex:0 0 auto}}
.donut svg{{transform:rotate(-90deg);width:100%;height:100%}}
.donut-track,.donut-segment{{fill:none;stroke-width:4.2}}.donut-track{{stroke:var(--border)}}
.donut-segment{{stroke-linecap:round;stroke-dasharray:0 100;stroke-dashoffset:var(--offset);animation:donut-draw 1.85s .5s cubic-bezier(.2,.8,.2,1) forwards}}
.donut-center{{position:absolute;inset:0;display:grid;place-content:center;text-align:center}}
.donut-center strong{{font-size:1.45rem;letter-spacing:-.04em}}.donut-center span{{margin-top:.1rem;color:var(--text3);font-size:.62rem;letter-spacing:.1em;text-transform:uppercase}}
.donut-legend{{display:grid;gap:.55rem;min-width:112px;font-size:.78rem}}
.legend-item{{display:grid;grid-template-columns:8px 1fr auto;align-items:center;gap:.55rem;color:var(--text2)}}
.legend-item b{{color:var(--text);font-family:var(--mono);font-size:.75rem}}.legend-dot{{width:7px;height:7px;border-radius:50%;display:block;background:var(--dot)}}
.outcome-distribution{{position:absolute;right:.2rem;bottom:-1px;left:.2rem;display:flex;gap:3px;height:4px;overflow:hidden;border-radius:99px;background:var(--border)}}
.outcome-distribution i{{display:block;min-width:0;transform:scaleX(0);transform-origin:left;border-radius:inherit;animation:distribution-in 1.25s cubic-bezier(.2,.8,.2,1) forwards}}
.outcome-distribution .dist-pass{{background:var(--green)}}.outcome-distribution .dist-fail{{background:var(--red);animation-delay:.12s}}.outcome-distribution .dist-amb{{background:var(--amber);animation-delay:.24s}}
@keyframes outcome-bar-down{{to{{height:72%}}}}@keyframes distribution-in{{to{{transform:scaleX(1)}}}}@keyframes donut-draw{{to{{stroke-dasharray:var(--dash)}}}}

/* Bottom row: active profile restrictions */
.header-bottom{{display:flex;align-items:flex-start}}
.restrictions{{display:flex;align-items:flex-start;flex-wrap:wrap;gap:1rem 2rem;width:100%}}
.criteria-section{{min-width:max-content}}
.criteria-section h4{{font-size:.78rem;text-transform:uppercase;letter-spacing:.05em;color:var(--text3);margin-bottom:.4rem}}
.chips{{display:flex;flex-wrap:wrap;gap:.4rem}}
.chip{{padding:.3rem .7rem;border:1px solid transparent;border-radius:12px;font-size:.8rem;font-weight:600}}
.chip-required{{background:var(--accent-g);color:var(--accent2);border-color:rgba(124,92,252,.16)}}
.chip-rule{{background:var(--bg3);color:var(--text);border-color:var(--border)}}
.chip-preferred{{background:var(--green-g);color:var(--green);border-color:rgba(34,197,94,.18)}}
.chip-excluded{{background:var(--red-g);color:var(--red);border-color:rgba(244,63,94,.18)}}
.chip-muted{{background:var(--bg3);color:var(--text2);border-color:var(--border)}}

/* Scroll-driven screening insights */
.insights-stage{{--demo-enter:0;--demo-reasons:0;--demo-scan:0;--demo-coverage:0;--demo-decision:0;--demo-morph:0;--demo-pass-glow:0;--demo-fail-glow:0;--demo-amb-glow:0;--demo-flight-x:0px;--demo-flight-y:0px;--demo-tilt-y:-10deg;--demo-tilt-x:4deg;--demo-depth:70px;--demo-orb-color:var(--accent);position:relative;height:350vh;border-bottom:1px solid var(--border)}}
.insights-sticky{{position:sticky;top:0;display:grid;grid-template-columns:minmax(330px,.72fr) minmax(610px,1.28fr);gap:clamp(1.25rem,3vw,3.5rem);align-items:center;height:100vh;min-height:680px;max-width:1600px;margin:0 auto;padding:clamp(1.5rem,3vw,3rem) clamp(1.5rem,3vw,3rem);overflow:hidden}}
.charts-column{{display:grid;gap:1rem;align-content:center;position:relative;z-index:8}}
.chart-card{{--chart-stage:1;position:relative;overflow:hidden;background:var(--bg2);border:1px solid var(--border);border-radius:var(--radius);padding:1.2rem 1.5rem;opacity:calc(.08 + var(--chart-stage) * .92);transform:translateY(calc((1 - var(--chart-stage)) * 30px)) scale(calc(.975 + var(--chart-stage) * .025));filter:blur(calc((1 - var(--chart-stage)) * 5px));transition:border-color .25s,box-shadow .25s;will-change:opacity,transform,filter}}
.chart-reasons{{--chart-stage:var(--demo-reasons)}}.chart-coverage{{--chart-stage:var(--demo-coverage)}}
.chart-card::before{{content:"";position:absolute;top:-1px;right:22%;left:22%;height:1px;background:linear-gradient(90deg,transparent,var(--accent),transparent);opacity:var(--chart-stage);box-shadow:0 0 12px var(--accent)}}
.chart-card h3{{font-size:.88rem;font-weight:600;margin-bottom:1rem;color:var(--text)}}
.chart-row{{display:flex;align-items:center;gap:.8rem;margin-bottom:.6rem}}
.chart-label{{font-size:.82rem;min-width:120px;color:var(--text2);white-space:nowrap;overflow:hidden;text-overflow:ellipsis}}
.chart-bar-track{{flex:1;height:10px;background:var(--bg4);border-radius:5px;overflow:hidden}}
.chart-bar-fill{{height:100%;border-radius:5px;transform:scaleX(var(--chart-stage));transform-origin:left;transition:transform .2s linear}}
.reason-fill{{background:var(--red)}}
.chart-value{{font-size:.82rem;font-weight:600;min-width:36px;text-align:right;font-family:var(--mono)}}
.podium-column{{position:relative;height:min(80vh,760px);min-height:620px;overflow:visible;perspective:1400px}}
.demo-heading{{position:absolute;z-index:20;top:.65rem;left:1.5rem;right:1.5rem;display:flex;justify-content:space-between;gap:1rem;align-items:flex-start}}
.demo-eyebrow{{display:block;color:var(--accent2);font-size:.66rem;font-weight:760;letter-spacing:.13em;text-transform:uppercase}}
.demo-heading h2{{margin-top:.35rem;font-size:clamp(1.25rem,2vw,1.8rem);letter-spacing:-.035em}}
.demo-heading p{{margin-top:.35rem;max-width:36ch;color:var(--text2);font-size:.72rem;line-height:1.45}}
.demo-progress{{width:128px;flex:0 0 auto;padding-top:.15rem}}
.demo-progress-copy{{display:flex;justify-content:space-between;color:var(--text3);font-family:var(--mono);font-size:.6rem;text-transform:uppercase;letter-spacing:.05em}}
.demo-progress-track{{height:3px;margin-top:.5rem;overflow:hidden;border-radius:99px;background:var(--border)}}
.demo-progress-bar{{display:block;width:100%;height:100%;transform:scaleX(0);transform-origin:left;background:linear-gradient(90deg,var(--accent),#4f9cff);box-shadow:0 0 10px var(--accent)}}
.demo-scene{{position:absolute;inset:0;transform-style:preserve-3d}}
.demo-cv-float{{--enter-y:calc((1 - var(--demo-enter)) * 70px);position:absolute;z-index:14;left:42%;top:43%;width:clamp(225px,20vw,292px);opacity:var(--demo-enter);transform:translate(-50%,-50%) translate3d(var(--demo-flight-x),calc(var(--enter-y) + var(--demo-flight-y)),var(--demo-depth)) rotateY(var(--demo-tilt-y)) rotateX(var(--demo-tilt-x));will-change:transform,opacity}}
.demo-cv{{position:relative;width:100%;aspect-ratio:.74;padding:18px;border:1px solid color-mix(in srgb,var(--accent) 38%,var(--border2));border-radius:calc(18px + var(--demo-morph) * 120px);overflow:hidden;background:linear-gradient(145deg,color-mix(in srgb,var(--bg4) 88%,var(--accent) 12%),var(--bg2));box-shadow:0 28px 60px rgba(0,0,0,.42),0 0 42px rgba(124,92,252,.13);opacity:calc(1 - var(--demo-morph));transform:scale(calc(1 - var(--demo-morph) * .86));transition:border-radius .12s linear;will-change:transform,opacity,border-radius}}
.demo-cv::before{{content:"";position:absolute;inset:0;background:linear-gradient(115deg,rgba(255,255,255,.08),transparent 22%,transparent 72%,rgba(124,92,252,.07));pointer-events:none}}
.demo-verdict-orb{{position:absolute;z-index:30;top:50%;left:50%;width:24px;height:24px;border:3px solid color-mix(in srgb,var(--demo-orb-color) 42%,white);border-radius:50%;background:var(--demo-orb-color);box-shadow:0 0 8px var(--demo-orb-color),0 0 20px color-mix(in srgb,var(--demo-orb-color) 68%,transparent),0 0 38px color-mix(in srgb,var(--demo-orb-color) 30%,transparent);opacity:var(--demo-morph);transform:translate(-50%,-50%) scale(calc(.35 + var(--demo-morph) * .65));transition:background .24s,border-color .24s,box-shadow .24s;will-change:transform,opacity}}
.demo-verdict-orb::after{{content:"";position:absolute;inset:-6px;border:1px solid color-mix(in srgb,var(--demo-orb-color) 55%,transparent);border-radius:50%;opacity:calc(.25 + var(--demo-morph) * .75);animation:demo-orb-breathe 1.8s ease-in-out infinite}}
@keyframes demo-orb-breathe{{50%{{transform:scale(1.18);opacity:.25}}}}
.demo-cv-header{{position:relative;display:grid;grid-template-columns:42px 1fr;align-items:center;gap:10px;padding-bottom:11px;border-bottom:1px solid var(--border)}}
.demo-cv-mark{{width:42px;height:42px;display:grid;place-items:center;border-radius:50%;background:var(--accent-g);border:1px solid rgba(124,92,252,.34);color:var(--accent2);font-weight:800}}
.demo-cv-header strong{{display:block;font-size:.78rem}}.demo-cv-header span{{display:block;margin-top:.2rem;color:var(--text2);font-size:.55rem}}
.demo-cv-section{{position:relative;margin-top:12px}}
.demo-cv-section h4{{margin-bottom:7px;color:var(--text2);font-size:.48rem;letter-spacing:.11em;text-transform:uppercase}}
.demo-line{{height:5px;margin:5px 0;border-radius:99px;background:linear-gradient(90deg,color-mix(in srgb,var(--text2) 50%,transparent),rgba(92,102,128,.14))}}.demo-line-long{{width:100%}}.demo-line-mid{{width:76%}}.demo-line-short{{width:54%}}
.demo-role-row{{display:grid;grid-template-columns:7px 1fr;gap:8px}}.demo-role-row::before{{content:"";width:5px;height:5px;margin-top:5px;border-radius:50%;background:var(--accent);box-shadow:0 0 8px var(--accent)}}
.demo-skill{{display:grid;grid-template-columns:67px 1fr 27px;align-items:center;gap:6px;margin:6px 0;color:var(--text2);font-size:.5rem}}
.demo-skill b{{height:5px;overflow:hidden;border-radius:99px;background:var(--border)}}.demo-skill b i{{display:block;width:var(--skill);height:100%;border-radius:inherit;background:linear-gradient(90deg,var(--accent),#4f9cff);transform:scaleX(var(--demo-scan));transform-origin:left}}.demo-skill em{{font-style:normal;text-align:right;color:var(--text)}}
.demo-scan-beam{{position:absolute;z-index:20;inset:0;transform:translateY(calc((var(--demo-scan) * 112%) - 7%));opacity:clamp(0,calc(var(--demo-scan) * 4),1);pointer-events:none}}
.demo-scan-line{{position:absolute;top:0;right:-14%;left:-14%;height:2px;background:#c5b6ff;box-shadow:0 0 5px #fff,0 0 18px #9b6dff,0 0 38px #6c4dff}}
.demo-scan-wash{{position:absolute;top:-34px;right:-8%;left:-8%;height:70px;background:linear-gradient(to bottom,transparent,rgba(137,94,255,.2),transparent);filter:blur(6px)}}
.demo-podium{{position:absolute;z-index:3;left:42%;bottom:5%;width:clamp(330px,36vw,500px);height:180px;transform:translateX(-50%) rotateX(61deg);transform-style:preserve-3d}}
.demo-podium::before{{content:"";position:absolute;left:50%;bottom:-24px;width:86%;height:84px;transform:translateX(-50%);border-radius:50%;background:radial-gradient(ellipse,rgba(124,92,252,.34),rgba(124,92,252,.12) 50%,transparent 73%);filter:blur(17px)}}
.demo-podium-top,.demo-podium-rim,.demo-podium-base{{position:absolute;left:50%;border-radius:50%;transform:translateX(-50%)}}
.demo-podium-top{{z-index:4;top:0;width:74%;aspect-ratio:1/.26;background:radial-gradient(ellipse at center,rgba(255,255,255,.22) 0 8%,color-mix(in srgb,var(--accent) 30%,var(--bg3)) 18%,var(--bg3) 42%,var(--bg) 100%);border:1px solid var(--border2);box-shadow:inset 0 2px 1px rgba(255,255,255,.18),inset 0 -18px 24px rgba(0,0,0,.48),0 0 40px rgba(124,92,252,.22)}}
.demo-podium-core{{position:absolute;inset:30%;border-radius:50%;background:radial-gradient(ellipse,#fff 0,var(--accent) 12%,transparent 64%);filter:blur(5px);opacity:calc(.24 + var(--demo-scan) * .5)}}
.demo-podium-rim{{z-index:3;top:23px;width:90%;aspect-ratio:1/.28;background:linear-gradient(180deg,var(--bg4),var(--bg2) 58%,var(--bg));border:1px solid var(--border2);box-shadow:0 0 0 6px var(--bg),0 0 0 7px rgba(124,92,252,.14),inset 0 8px 12px rgba(255,255,255,.05),inset 0 -14px 18px rgba(0,0,0,.55),0 20px 34px rgba(0,0,0,.35)}}
.demo-podium-base{{z-index:2;top:66px;width:98%;aspect-ratio:1/.23;background:linear-gradient(180deg,var(--bg3),var(--bg2) 42%,var(--bg));border:1px solid var(--border);box-shadow:inset 0 10px 12px rgba(255,255,255,.025),inset 0 -20px 20px rgba(0,0,0,.5),0 22px 34px rgba(0,0,0,.5)}}
.demo-podium-base::before{{content:"";position:absolute;top:44%;right:8%;left:8%;height:20%;border-radius:99px;background:linear-gradient(90deg,transparent,var(--accent),transparent);filter:blur(6px);opacity:.72}}
.demo-podium-reflection{{position:absolute;top:108px;left:50%;width:82%;height:108px;transform:translateX(-50%);background:radial-gradient(ellipse,rgba(124,92,252,.25),transparent 70%);filter:blur(22px)}}
.demo-cv-shadow{{position:absolute;left:42%;bottom:20%;width:190px;height:42px;transform:translateX(-50%) scale(calc(1 - var(--demo-morph) * .72));border-radius:50%;background:rgba(0,0,0,.62);filter:blur(15px);opacity:calc(.58 - var(--demo-morph) * .5)}}
.demo-outcomes{{position:absolute;z-index:12;top:25%;right:1.1rem;display:grid;gap:10px;width:clamp(155px,14vw,205px)}}
.demo-outcome{{--outcome-color:var(--text2);--outcome-glow:0;display:grid;grid-template-columns:38px 1fr;gap:9px;align-items:center;padding:11px;border:1px solid color-mix(in srgb,var(--outcome-color) 24%,var(--border));border-radius:14px;background:color-mix(in srgb,var(--bg2) 88%,transparent);backdrop-filter:blur(12px);opacity:calc(.14 + var(--demo-decision) * .86);transform:translateX(calc((1 - var(--demo-decision)) * 20px)) scale(calc(1 + var(--outcome-glow) * .035));filter:saturate(calc(.35 + var(--demo-decision) * .65));box-shadow:0 0 calc(var(--outcome-glow) * 44px) color-mix(in srgb,var(--outcome-color) 58%,transparent);transition:border-color .18s,background .18s,box-shadow .18s,transform .18s}}
.demo-outcome-pass{{--outcome-color:var(--green);--outcome-glow:var(--demo-pass-glow)}}
.demo-outcome-fail{{--outcome-color:var(--red);--outcome-glow:var(--demo-fail-glow)}}
.demo-outcome-amb{{--outcome-color:var(--amber);--outcome-glow:var(--demo-amb-glow)}}
.demo-outcome-icon{{width:36px;height:36px;display:grid;place-items:center;border-radius:50%;background:color-mix(in srgb,var(--outcome-color) 18%,var(--bg3));color:var(--outcome-color);font-size:1rem;font-weight:800;box-shadow:0 0 18px color-mix(in srgb,var(--outcome-color) 18%,transparent)}}
.demo-outcome span{{display:block;color:var(--text2);font-size:.53rem;font-weight:700;letter-spacing:.08em;text-transform:uppercase}}.demo-outcome strong{{display:block;margin-top:.15rem;color:var(--outcome-color);font-family:var(--mono);font-size:.82rem}}
.demo-caption{{position:absolute;right:1.3rem;bottom:1.1rem;color:var(--text3);font-size:.6rem}}

/* Main area */
.main{{padding:1.5rem 2rem;max-width:1400px;margin:0 auto;width:100%}}
.criterion-lab{{margin:1rem 0;overflow:hidden;border:1px solid color-mix(in srgb,var(--accent) 36%,var(--border));border-radius:var(--radius);background:linear-gradient(145deg,color-mix(in srgb,var(--bg2) 94%,var(--accent) 6%),var(--bg2));box-shadow:0 18px 45px rgba(0,0,0,.1)}}
.criterion-lab-head{{display:flex;align-items:flex-start;justify-content:space-between;gap:1.5rem;padding:1.2rem 1.25rem;border-bottom:1px solid var(--border)}}
.criterion-lab-kicker{{display:block;margin-bottom:.28rem;color:var(--accent2);font-size:.64rem;font-weight:780;letter-spacing:.12em;text-transform:uppercase}}
.criterion-lab-head h2{{font-size:1.2rem;letter-spacing:-.025em}}
.criterion-lab-head p{{max-width:72ch;margin-top:.38rem;color:var(--text2);font-size:.76rem;line-height:1.5}}
.criterion-lab-safety{{display:inline-flex;align-items:center;gap:.38rem;flex:0 0 auto;padding:.35rem .6rem;border:1px solid rgba(34,197,94,.22);border-radius:999px;background:var(--green-g);color:var(--green);font-size:.65rem;font-weight:720;white-space:nowrap}}
.criterion-lab-safety::before{{content:"";width:6px;height:6px;border-radius:50%;background:currentColor;box-shadow:0 0 8px currentColor}}
.criterion-lab-grid{{display:grid;grid-template-columns:minmax(360px,.92fr) minmax(430px,1.08fr)}}
.criterion-lab-controls,.criterion-lab-results{{padding:1.2rem 1.25rem}}
.criterion-lab-controls{{border-right:1px solid var(--border)}}
.criterion-panel-title{{display:flex;align-items:center;justify-content:space-between;gap:1rem;margin-bottom:.9rem}}
.criterion-panel-title h3{{font-size:.83rem;font-weight:720}}
.criterion-panel-title span{{color:var(--text3);font-size:.66rem}}
.criterion-years{{display:grid;grid-template-columns:1fr auto;gap:.7rem 1rem;align-items:center;padding:.85rem;border:1px solid var(--border);border-radius:10px;background:var(--bg3)}}
.criterion-years-copy strong{{display:block;font-size:.8rem}}.criterion-years-copy span{{display:block;margin-top:.2rem;color:var(--text3);font-size:.66rem}}
.criterion-years-output{{min-width:68px;color:var(--accent2);font-family:var(--mono);font-size:1.05rem;font-weight:760;text-align:right}}
.criterion-years input{{grid-column:1/-1;width:100%;accent-color:var(--accent);cursor:pointer}}
.criterion-rules{{display:grid;gap:.55rem;margin-top:.75rem}}
.criterion-rule{{display:flex;align-items:center;justify-content:space-between;gap:.8rem;padding:.58rem .65rem;border:1px solid var(--border);border-radius:10px;background:var(--bg3)}}
.criterion-rule-name{{display:flex;align-items:center;gap:.55rem;min-width:0;font-size:.76rem;font-weight:680}}
.criterion-rule-name .kw-tool-icon{{width:30px;height:30px;flex-basis:30px;border-radius:8px}}
.criterion-rule-name .kw-tool-icon img{{width:22px;height:22px}}
.criterion-mode-group{{display:flex;overflow:hidden;border:1px solid var(--border2);border-radius:8px;background:var(--bg)}}
.criterion-mode-btn{{padding:.38rem .48rem;border:0;border-left:1px solid var(--border);background:transparent;color:var(--text3);font:inherit;font-size:.6rem;font-weight:680;cursor:pointer;transition:background .15s,color .15s}}
.criterion-mode-btn:first-child{{border-left:0}}
.criterion-mode-btn:hover{{color:var(--text)}}
.criterion-mode-btn.is-active[data-mode="required"]{{background:var(--accent-g);color:var(--accent2)}}
.criterion-mode-btn.is-active[data-mode="preferred"]{{background:var(--green-g);color:var(--green)}}
.criterion-mode-btn.is-active[data-mode="off"]{{background:var(--bg4);color:var(--text2)}}
.criterion-lab-actions{{display:flex;flex-wrap:wrap;align-items:center;gap:.55rem;margin-top:.85rem}}
.criterion-reset{{padding:.48rem .72rem;border:1px solid var(--border2);border-radius:8px;background:var(--bg3);color:var(--text2);font:inherit;font-size:.68rem;font-weight:680;cursor:pointer}}
.criterion-reset:hover{{border-color:var(--accent);color:var(--text)}}
.criterion-apply{{padding:.5rem .82rem;border:1px solid color-mix(in srgb,var(--accent) 65%,var(--border));border-radius:8px;background:linear-gradient(135deg,var(--accent),#4f7cff);color:#fff;font:inherit;font-size:.68rem;font-weight:760;cursor:pointer;box-shadow:0 5px 16px rgba(124,92,252,.2);transition:transform .15s,opacity .15s}}
.criterion-apply:hover:not(:disabled){{transform:translateY(-1px)}}
.criterion-apply:disabled{{opacity:.42;cursor:not-allowed;box-shadow:none}}
.criterion-adjustment-count{{margin-left:auto;color:var(--text3);font-size:.66rem}}
.criterion-apply-status{{width:100%;min-height:1em;color:var(--text3);font-size:.62rem;line-height:1.4}}
.criterion-apply-status.is-applied{{color:var(--green)}}
.criterion-result-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:.55rem}}
.criterion-result{{padding:.75rem;border:1px solid var(--border);border-radius:10px;background:var(--bg3)}}
.criterion-result span{{display:block;color:var(--text3);font-size:.58rem;font-weight:720;letter-spacing:.07em;text-transform:uppercase}}
.criterion-result strong{{display:block;margin-top:.25rem;font-family:var(--mono);font-size:1.22rem}}
.criterion-result-pass strong{{color:var(--green)}}.criterion-result-fail strong{{color:var(--red)}}.criterion-result-amb strong{{color:var(--amber)}}.criterion-result-pref strong{{color:var(--accent2)}}
.criterion-result small{{display:block;min-height:1em;margin-top:.18rem;color:var(--text3);font-size:.58rem}}
.criterion-distribution{{display:flex;gap:3px;height:6px;margin:.7rem 0 1rem;overflow:hidden;border-radius:999px;background:var(--border)}}
.criterion-distribution i{{display:block;min-width:0;border-radius:inherit;transition:flex .25s ease}}
.criterion-distribution-pass{{background:var(--green)}}.criterion-distribution-fail{{background:var(--red)}}.criterion-distribution-amb{{background:var(--amber)}}
.criterion-lab-insights{{display:grid;grid-template-columns:1fr 1fr;gap:.75rem}}
.criterion-subpanel{{min-width:0;padding:.8rem;border:1px solid var(--border);border-radius:10px;background:var(--bg3)}}
.criterion-subpanel h3{{font-size:.72rem;font-weight:720}}
.criterion-impact-list{{display:grid;gap:.62rem;margin-top:.72rem}}
.criterion-impact-row>div:first-child{{display:flex;align-items:flex-start;justify-content:space-between;gap:.6rem}}
.criterion-impact-row strong{{font-size:.65rem;font-weight:650}}
.criterion-impact-copy{{color:var(--text3);font-size:.58rem;text-align:right}}
.criterion-impact-track{{height:4px;margin-top:.32rem;overflow:hidden;border-radius:99px;background:var(--border)}}
.criterion-impact-track i{{display:block;width:0;height:100%;border-radius:inherit;background:linear-gradient(90deg,var(--accent),#4f9cff);transition:width .25s ease}}
.criterion-warning-list,.criterion-movement-list{{display:grid;gap:.42rem;margin-top:.7rem}}
.criterion-warning{{padding:.5rem .58rem;border-left:2px solid var(--amber);border-radius:6px;background:var(--amber-g);color:var(--text2);font-size:.62rem;line-height:1.4}}
.criterion-warning.is-positive{{border-left-color:var(--green);background:var(--green-g)}}
.criterion-movement{{display:grid;grid-template-columns:minmax(0,1fr) auto;gap:.5rem;align-items:center;padding:.42rem .5rem;border-radius:6px;background:var(--bg)}}
.criterion-movement-name{{overflow:hidden;color:var(--text2);font-size:.61rem;text-overflow:ellipsis;white-space:nowrap}}
.criterion-movement-change{{font-family:var(--mono);font-size:.58rem;font-weight:700}}
.criterion-movement-more,.criterion-empty{{color:var(--text3);font-size:.6rem;line-height:1.4}}
.criterion-lab-foot{{padding:.72rem 1.25rem;border-top:1px solid var(--border);background:color-mix(in srgb,var(--bg3) 75%,transparent);color:var(--text3);font-size:.64rem;line-height:1.45}}
.lab-application-note{{margin-bottom:.75rem;padding:.62rem .72rem;border:1px solid color-mix(in srgb,var(--accent) 30%,var(--border));border-left:3px solid var(--accent);border-radius:8px;background:var(--accent-g);color:var(--text2);font-size:.7rem;line-height:1.45}}
.tool-filter-panel{{margin-bottom:1rem;padding:1rem;background:var(--bg2);border:1px solid var(--border);border-radius:var(--radius)}}
.tool-filter-head{{display:flex;align-items:flex-start;justify-content:space-between;gap:1rem;margin-bottom:.9rem}}
.tool-filter-kicker{{display:block;margin-bottom:.25rem;color:var(--accent2);font-size:.64rem;font-weight:760;letter-spacing:.11em;text-transform:uppercase}}
.tool-filter-head h2{{font-size:1rem;letter-spacing:-.015em}}
.tool-filter-head p{{margin-top:.3rem;color:var(--text3);font-size:.73rem}}
.tool-filter-clear{{padding:.42rem .7rem;border:1px solid var(--border2);border-radius:8px;background:var(--bg3);color:var(--text2);font-size:.72rem;font-weight:650;cursor:pointer}}
.tool-filter-clear:disabled{{opacity:.4;cursor:not-allowed}}
.tool-filter-chips{{display:flex;flex-wrap:wrap;gap:.5rem}}
.tool-filter-chip{{display:inline-flex;align-items:center;gap:.42rem;padding:.42rem .58rem;border:1px solid var(--border);border-radius:9px;background:var(--bg3);color:var(--text2);font:inherit;font-size:.75rem;font-weight:650;cursor:pointer;transition:border-color .15s,background .15s,color .15s,transform .15s}}
.tool-filter-chip:hover{{transform:translateY(-1px);border-color:var(--accent);color:var(--text)}}
.tool-filter-chip[aria-pressed="true"]{{border-color:var(--accent);background:var(--accent-g);color:var(--accent2);box-shadow:inset 0 0 0 1px color-mix(in srgb,var(--accent) 24%,transparent)}}
.tool-filter-chip img{{width:18px;height:18px;object-fit:contain}}
.tool-filter-chip img.tool-icon-enlarged{{width:24px;height:24px;object-fit:fill}}
.tool-filter-chip img.tool-icon-docker{{width:28px;height:16px;object-fit:contain}}
.tool-filter-chip img.tool-icon-zabbix{{transform:scaleY(1.65)}}
.tool-filter-chip strong{{min-width:1.4rem;padding:.08rem .3rem;border-radius:6px;background:var(--bg);color:var(--text3);font-family:var(--mono);font-size:.65rem;text-align:center}}
.tool-filter-chip[aria-pressed="true"] strong{{color:var(--accent2)}}
.tool-filter-summary{{margin-top:.8rem;color:var(--text3);font-size:.72rem}}

/* Table */
.table-wrap{{background:var(--bg2);border:1px solid var(--border);border-radius:var(--radius);overflow:hidden}}
table{{width:100%;border-collapse:collapse;font-size:.95rem}}
thead th{{text-align:left;padding:.7rem .8rem;font-size:.78rem;text-transform:uppercase;letter-spacing:.05em;color:var(--text3);border-bottom:1px solid var(--border);background:var(--bg3)}}
tbody tr.data-row{{cursor:pointer;transition:background .15s}}
tbody tr.data-row:hover{{background:var(--bg4)}}
tbody tr.data-row.selected{{background:var(--accent-g);border-left:3px solid var(--accent)}}
tbody td{{padding:.65rem .8rem;border-bottom:1px solid var(--border);vertical-align:middle}}
.cell-name{{font-weight:500;max-width:260px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}

/* Score bar */
.score-bar-cell{{display:flex;align-items:center;gap:.5rem}}
.score-bar-track{{flex:1;height:4px;background:var(--border);border-radius:2px;overflow:hidden;min-width:40px}}
.score-bar-fill{{height:100%;border-radius:2px;transition:width .6s ease}}
.score-fill-pass{{background:var(--green)}}
.score-fill-fail{{background:var(--red)}}
.score-fill-ambiguous{{background:var(--amber)}}
.score-num{{font-family:var(--mono);font-size:.85rem;font-weight:600;min-width:22px}}

/* Pills */
.pill{{display:inline-block;padding:.15rem .5rem;border-radius:10px;font-size:.72rem;font-weight:600;text-transform:uppercase;letter-spacing:.03em}}
.pill-pass{{background:var(--green-g);color:var(--green)}}
.pill-fail{{background:var(--red-g);color:var(--red)}}
.pill-ambiguous{{background:var(--amber-g);color:var(--amber)}}

/* Inline candidate review */
.detail-row td{{padding:0;border-bottom:1px solid var(--border)}}
.candidate-review{{padding:1rem 1.15rem 1.35rem;background:var(--bg);animation:reviewIn .2s ease}}
@keyframes reviewIn{{from{{opacity:0;transform:translateY(-7px)}}to{{opacity:1;transform:translateY(0)}}}}

/* Candidate review workspace */
.review-hero{{display:flex;align-items:center;justify-content:space-between;gap:1.5rem;padding:1.15rem 1.25rem;background:linear-gradient(135deg,var(--bg2),var(--bg3));border:1px solid var(--border);border-top:3px solid var(--accent);border-radius:12px 12px 0 0}}
.review-hero-pass{{border-top-color:var(--green)}}.review-hero-fail{{border-top-color:var(--red)}}.review-hero-ambiguous{{border-top-color:var(--amber)}}
.review-outcome{{display:flex;align-items:center;gap:.9rem;min-width:0}}
.outcome-icon{{width:44px;height:44px;border-radius:12px;display:grid;place-items:center;background:var(--bg4);font-size:1.25rem;font-weight:800;flex:0 0 auto}}
.review-hero-pass .outcome-icon{{background:var(--green-g);color:var(--green)}}.review-hero-fail .outcome-icon{{background:var(--red-g);color:var(--red)}}.review-hero-ambiguous .outcome-icon{{background:var(--amber-g);color:var(--amber)}}
.review-eyebrow{{font-size:.66rem;font-weight:700;text-transform:uppercase;letter-spacing:.12em;color:var(--text3);margin-bottom:.2rem}}
.review-outcome h3{{font-size:1.35rem;line-height:1;font-weight:780;letter-spacing:-.02em}}
.review-hero-pass h3{{color:var(--green)}}.review-hero-fail h3{{color:var(--red)}}.review-hero-ambiguous h3{{color:var(--amber)}}
.review-outcome p{{font-size:.8rem;color:var(--text2);margin-top:.35rem}}
.review-metrics{{display:flex;align-items:stretch;border:1px solid var(--border);border-radius:10px;background:var(--bg2);overflow:hidden;flex:0 0 auto}}
.review-metric{{min-width:82px;padding:.62rem .8rem;text-align:center;border-left:1px solid var(--border)}}
.review-metric:first-child{{border-left:0}}
.review-metric strong{{display:block;font-size:1.05rem;font-family:var(--mono);line-height:1.2}}
.review-metric span{{display:block;margin-top:.2rem;font-size:.62rem;text-transform:uppercase;letter-spacing:.07em;color:var(--text3)}}
.review-body{{display:flex;flex-direction:column;gap:.85rem;padding:1rem;background:var(--bg3);border:1px solid var(--border);border-top:0;border-radius:0 0 12px 12px}}
.decision-rationale{{display:grid;grid-template-columns:minmax(180px,.32fr) 1fr;gap:1rem;align-items:start;padding:.9rem 1rem;background:var(--bg2);border:1px solid var(--border);border-left:3px solid var(--accent);border-radius:10px}}
.rationale-pass{{border-left-color:var(--green)}}.rationale-fail{{border-left-color:var(--red)}}.rationale-ambiguous{{border-left-color:var(--amber)}}
.rationale-heading span{{display:block;font-size:.86rem;font-weight:700}}
.rationale-heading small{{display:block;color:var(--text3);font-size:.68rem;margin-top:.2rem}}
.decision-rationale ul{{margin:0;padding-left:1.1rem;color:var(--text2);font-size:.78rem;line-height:1.55}}
.review-section{{background:var(--bg2);border:1px solid var(--border);border-radius:10px;padding:1rem}}
.review-section-head{{display:flex;align-items:center;gap:.7rem;padding-bottom:.8rem;margin-bottom:.8rem;border-bottom:1px solid var(--border)}}
.section-icon{{width:30px;height:30px;display:grid;place-items:center;border-radius:8px;background:var(--accent-g);color:var(--accent2);font-size:.65rem;font-weight:800;letter-spacing:.04em;flex:0 0 auto}}
.section-icon-ai{{width:40px;height:40px;border-radius:0;background:transparent;color:var(--amber)}}
.section-icon-ai .copilot-icon{{display:block;width:38px;height:38px;object-fit:contain}}
.section-icon-codex{{width:38px;height:38px;border:1px solid color-mix(in srgb,var(--accent) 42%,var(--border));border-radius:50%;background:radial-gradient(circle at 35% 30%,#a78bfa,var(--accent) 45%,#312e81);color:#fff;font-family:var(--mono);font-size:.67rem;font-weight:850;letter-spacing:-.03em;box-shadow:0 0 0 4px var(--accent-g),0 7px 18px rgba(124,92,252,.2)}}
.review-section-head h4{{font-size:.9rem;font-weight:700;line-height:1.2}}
.review-section-head p{{font-size:.7rem;color:var(--text3);margin-top:.2rem}}
.keyword-grid{{display:flex;flex-direction:column;gap:1rem}}
.review-split{{display:grid;grid-template-columns:1.15fr .85fr;gap:.85rem;align-items:stretch}}

/* KW items */
.kw-item{{width:100%;padding:1.05rem 1.15rem;border-radius:12px;border:1px solid var(--border);background:var(--bg3);min-width:0}}
.kw-found{{border-top:2px solid var(--green)}}.kw-missing{{border-top:2px solid var(--red)}}.kw-review{{border-top:2px solid var(--amber)}}
.kw-header{{display:flex;justify-content:space-between;align-items:center;gap:1rem;padding-bottom:.85rem;border-bottom:1px solid var(--border)}}
.kw-title{{display:flex;align-items:center;gap:.75rem;min-width:0}}
.kw-tool-icon{{width:44px;height:44px;display:grid;place-items:center;flex:0 0 44px;border-radius:11px;background:var(--bg);border:1px solid var(--border);box-shadow:0 6px 18px rgba(15,23,42,.08)}}
.kw-tool-icon img{{display:block;width:30px;height:30px;object-fit:contain}}
.kw-tool-icon img.tool-icon-enlarged{{width:38px;height:38px;object-fit:fill}}
.kw-tool-icon img.tool-icon-docker{{width:38px;height:22px;object-fit:contain}}
.kw-tool-icon img.tool-icon-zabbix{{transform:scaleY(1.65)}}
.tool-icon-dark{{display:none!important}}
[data-theme="dark"] .tool-icon-light{{display:none!important}}
[data-theme="dark"] .tool-icon-dark{{display:block!important}}
[data-theme="light"] .tool-icon-light{{display:block!important}}
[data-theme="light"] .tool-icon-dark{{display:none!important}}
.kw-tool-icon-fallback{{font-family:var(--mono);font-size:.7rem;font-weight:800;color:var(--accent2);background:var(--accent-g)}}
.kw-title-copy{{display:flex;flex-direction:column;gap:.38rem}}
.kw-name{{font-weight:750;font-size:1.05rem;line-height:1.15;text-transform:capitalize}}
.kw-kicker{{font-size:.66rem;line-height:1.2;color:var(--text3);text-transform:uppercase;letter-spacing:.08em}}
.kw-badge{{font-size:.68rem;font-weight:600;text-transform:uppercase;padding:.1rem .4rem;border-radius:6px}}
.kw-badge.found{{background:var(--green-g);color:var(--green)}}.kw-badge.missing{{background:var(--red-g);color:var(--red)}}.kw-badge.review{{background:var(--amber-g);color:var(--amber)}}
.kw-relation{{font-size:.72rem;color:var(--text2);margin-top:.25rem}}
.kw-citations{{display:flex;flex-direction:column;gap:.8rem;margin-top:.9rem}}
.kw-citation{{width:100%;padding:.9rem 1rem;border:1px solid var(--border);border-left:3px solid var(--green);border-radius:10px;background:var(--bg2)}}
.kw-citation-review{{margin-top:.9rem;border-left-color:var(--amber)}}
.kw-citation-head{{display:flex;align-items:center;justify-content:space-between;gap:.75rem;margin-bottom:.6rem}}
.kw-citation-source{{display:flex;align-items:center;justify-content:flex-end;gap:.55rem;flex-wrap:wrap}}
.kw-citation-index{{font-size:.68rem;font-weight:750;text-transform:uppercase;letter-spacing:.07em;color:var(--green)}}
.kw-citation-review .kw-citation-index{{color:var(--amber)}}
.kw-citation .kw-relation{{margin-top:0;font-weight:600;text-align:right}}
.kw-snippet{{font-family:var(--mono);font-size:.78rem;background:var(--bg);padding:.85rem 1rem;border:1px solid var(--border);border-radius:8px;margin:0;white-space:pre-wrap;word-break:break-word;line-height:1.65;color:var(--text2);overflow:visible}}
.kw-hl{{background:#22c55e33;color:#16a34a;font-weight:600;padding:1px 3px;border-radius:3px}}
.xray-open-btn{{display:inline-flex;align-items:center;gap:.38rem;margin-top:.65rem;padding:.42rem .62rem;border:1px solid color-mix(in srgb,var(--accent) 45%,var(--border));border-radius:8px;background:var(--accent-g);color:var(--accent2);font:inherit;font-size:.66rem;font-weight:720;cursor:pointer;transition:border-color .15s,background .15s,transform .15s}}
.kw-citation-source .xray-open-btn{{margin-top:0}}
.xray-open-btn:hover{{transform:translateY(-1px);border-color:var(--accent);background:color-mix(in srgb,var(--accent-g) 70%,var(--bg3))}}

/* Evidence X-Ray */
body.xray-open{{overflow:hidden}}
.xray-shell{{position:fixed;z-index:1000;inset:0;display:grid;place-items:center;padding:2vh 2vw}}
.xray-shell[hidden]{{display:none}}
.xray-backdrop{{position:absolute;inset:0;border:0;background:rgba(3,6,13,.78);backdrop-filter:blur(8px)}}
.xray-dialog{{position:relative;display:flex;flex-direction:column;width:min(1500px,96vw);height:min(920px,96vh);overflow:hidden;border:1px solid var(--border2);border-radius:18px;background:var(--bg2);box-shadow:0 30px 90px rgba(0,0,0,.55)}}
.xray-head{{display:flex;align-items:center;justify-content:space-between;gap:1rem;padding:.9rem 1rem;border-bottom:1px solid var(--border);background:linear-gradient(135deg,color-mix(in srgb,var(--bg3) 92%,var(--accent) 8%),var(--bg2))}}
.xray-title-group{{display:flex;align-items:center;gap:.75rem;min-width:0}}
.xray-mark{{width:38px;height:38px;display:grid;place-items:center;flex:0 0 auto;border:1px solid rgba(124,92,252,.3);border-radius:10px;background:var(--accent-g);color:var(--accent2);font-size:1.05rem}}
.xray-kicker{{display:block;color:var(--accent2);font-size:.58rem;font-weight:760;letter-spacing:.12em;text-transform:uppercase}}
.xray-head h2{{margin-top:.15rem;font-size:1rem;letter-spacing:-.015em}}
.xray-candidate{{display:block;max-width:70vw;margin-top:.15rem;overflow:hidden;color:var(--text3);font-size:.65rem;text-overflow:ellipsis;white-space:nowrap}}
.xray-head-actions{{display:flex;align-items:center;gap:.45rem}}
.xray-nav,.xray-close{{height:34px;border:1px solid var(--border2);border-radius:8px;background:var(--bg3);color:var(--text2);font:inherit;cursor:pointer}}
.xray-nav{{min-width:34px;padding:0 .55rem;font-size:.68rem;font-weight:680}}
.xray-close{{width:34px;font-size:1rem}}
.xray-nav:hover,.xray-close:hover{{border-color:var(--accent);color:var(--text)}}
.xray-nav:disabled{{opacity:.35;cursor:not-allowed}}
.xray-body{{display:grid;grid-template-columns:minmax(0,1.48fr) minmax(330px,.72fr);min-height:0;flex:1}}
.xray-source{{display:flex;flex-direction:column;min-width:0;min-height:0;border-right:1px solid var(--border);background:#161922}}
.xray-source-head{{display:flex;align-items:center;justify-content:space-between;gap:1rem;padding:.62rem .8rem;border-bottom:1px solid var(--border);background:var(--bg3)}}
.xray-source-head span{{font-size:.68rem;font-weight:680}}
.xray-page{{color:var(--text3);font-family:var(--mono);font-size:.62rem}}
.xray-preview{{position:relative;display:flex;align-items:flex-start;justify-content:center;min-height:0;flex:1;padding:1rem;overflow:auto;background:#20232d}}
.xray-page-image{{display:block;max-width:100%;height:auto;background:#fff;box-shadow:0 12px 38px rgba(0,0,0,.38);transition:opacity .18s}}
.xray-page-image.is-loading{{opacity:.18}}
.xray-loading{{position:absolute;z-index:2;top:1.6rem;left:50%;transform:translateX(-50%);padding:.45rem .65rem;border:1px solid var(--border2);border-radius:8px;background:var(--bg2);color:var(--text2);font-size:.65rem;white-space:nowrap;box-shadow:0 8px 24px rgba(0,0,0,.24)}}
.xray-source-fallback{{display:grid;place-content:center;gap:.7rem;height:100%;padding:2rem;text-align:center;color:var(--text2)}}
.xray-preview[hidden],.xray-loading[hidden],.xray-source-fallback[hidden],.xray-open-original[hidden]{{display:none}}
.xray-source-fallback strong{{font-size:.9rem}}.xray-source-fallback span{{max-width:48ch;font-size:.7rem;line-height:1.55}}
.xray-open-original{{display:inline-flex;align-items:center;justify-content:center;padding:.5rem .7rem;border:1px solid var(--border2);border-radius:8px;background:var(--bg3);color:var(--text);font-size:.68rem;font-weight:680;text-decoration:none}}
.xray-open-original:hover{{border-color:var(--accent);color:var(--accent2)}}
.xray-evidence{{min-height:0;padding:1rem;overflow:auto;background:var(--bg2)}}
.xray-badges{{display:flex;flex-wrap:wrap;gap:.4rem}}
.xray-badge{{padding:.25rem .48rem;border:1px solid var(--border);border-radius:999px;background:var(--bg3);color:var(--text2);font-size:.6rem;font-weight:680;text-transform:capitalize}}
.xray-badge-primary{{border-color:rgba(34,197,94,.24);background:var(--green-g);color:var(--green)}}
.xray-facts{{display:grid;gap:.55rem;margin-top:.9rem;padding:.75rem;border:1px solid var(--border);border-radius:10px;background:var(--bg3)}}
.xray-fact{{display:grid;grid-template-columns:82px 1fr;gap:.6rem;font-size:.66rem;line-height:1.4}}
.xray-fact span{{color:var(--text3)}}.xray-fact strong{{min-width:0;font-weight:650;word-break:break-word}}
.xray-evidence h3{{margin-top:1rem;font-size:.72rem}}
.xray-excerpt{{margin-top:.55rem;padding:.85rem;border:1px solid var(--border);border-left:3px solid var(--green);border-radius:8px;background:var(--bg);color:var(--text2);font-family:var(--mono);font-size:.68rem;line-height:1.65;white-space:pre-wrap;word-break:break-word}}
.xray-excerpt.is-updated{{animation:xray-excerpt-update .28s ease}}
.xray-excerpt mark{{padding:1px 3px;border-radius:3px;background:#facc15;color:#111827;font-weight:800;box-shadow:0 0 0 1px rgba(250,204,21,.35),0 0 14px rgba(250,204,21,.18)}}
@keyframes xray-excerpt-update{{from{{transform:translateY(3px);opacity:.55}}to{{transform:none;opacity:1}}}}
.xray-note{{margin-top:.75rem;padding:.62rem .7rem;border-radius:8px;background:var(--accent-g);color:var(--text2);font-size:.62rem;line-height:1.5}}

/* Timeline */
.timeline{{--exp-axis:4px;--exp-gap:.8rem;display:flex;flex-direction:column;position:relative;margin-left:.35rem}}
.exp-item{{position:relative;padding:.25rem 0 .25rem 1rem}}
.exp-item-connected{{margin-bottom:var(--exp-gap)}}
.exp-dot{{position:absolute;z-index:2;left:var(--exp-axis);top:.48rem;width:9px;height:9px;border-radius:50%;background:var(--accent);border:2px solid var(--bg2);box-shadow:0 0 0 1px var(--accent);transform:translateX(-50%)}}
.exp-content{{position:relative;z-index:1}}
.exp-title{{font-weight:600;font-size:.9rem}}
.exp-meta{{font-size:.78rem;color:var(--text2);margin-top:.1rem}}
.exp-connector{{position:absolute;z-index:0;top:calc(.48rem + 4.5px);left:0;width:100%;height:calc(100% + var(--exp-gap));overflow:hidden;pointer-events:none}}
.exp-connector::before{{position:absolute;top:0;bottom:0;left:var(--exp-axis);width:1px;background:var(--border2);content:"";transform:translateX(-50%)}}
.exp-particle{{--exp-particle-half:2px;position:absolute;z-index:1;top:calc(100% + var(--exp-particle-half));left:var(--exp-axis);width:4px;height:4px;border-radius:50%;background:var(--accent);filter:drop-shadow(0 0 4px var(--accent));opacity:0;will-change:top,opacity;animation:exp-flow-up 4.2s linear infinite}}
.exp-particle:nth-child(2){{animation-delay:-1.4s}}
.exp-particle:nth-child(3){{animation-delay:-2.8s}}
@keyframes exp-flow-up{{
    0%{{top:calc(100% + var(--exp-particle-half));transform:translateX(-50%);opacity:0}}
    10%{{opacity:1}}
    96%{{opacity:1}}
    100%{{top:-4px;transform:translateX(-50%);opacity:0}}
}}

/* Flags */
.flag-item{{display:flex;align-items:flex-start;gap:.5rem;padding:.58rem .65rem;border-radius:var(--radius-sm);margin-bottom:.4rem;font-size:.8rem;line-height:1.4}}
.flag-info{{background:var(--bg4)}}.flag-danger{{background:var(--red-g)}}.flag-success{{background:var(--green-g)}}
.flag-icon{{font-size:.9rem}}
.verdict-item{{margin-bottom:.18rem}}

.empty-state{{color:var(--text3);font-size:.8rem;font-style:italic;padding:.5rem 0}}
.file-link{{text-decoration:none;border-bottom:1px dashed var(--text3);transition:all .15s}}
.file-link:hover{{color:var(--accent);border-color:var(--accent)}}

/* Responsive */
@media(max-width:1100px){{
    .header-middle{{grid-template-columns:minmax(185px,.5fr) 1.5fr}}
    .donut-wrap{{grid-column:1/-1;min-height:150px;padding:1rem 0 0;border-top:1px solid var(--border);border-left:0}}
    .insights-sticky{{grid-template-columns:minmax(300px,.68fr) minmax(520px,1.32fr);padding-right:1.25rem;padding-left:1.25rem}}
    .demo-outcomes{{right:.65rem;width:155px}}.demo-outcome{{grid-template-columns:31px 1fr;padding:9px}}.demo-outcome-icon{{width:30px;height:30px}}
    .criterion-lab-grid{{grid-template-columns:1fr}}
    .criterion-lab-controls{{border-right:0;border-bottom:1px solid var(--border)}}
}}
@media(max-width:900px){{
    .insights-stage{{height:auto}}
    .insights-sticky{{position:relative;grid-template-columns:1fr;height:auto;min-height:0;padding:1.25rem;overflow:visible}}
    .charts-column{{order:1}}.podium-column{{order:2;height:650px;min-height:0}}
    .demo-cv-float{{left:44%;width:min(39vw,270px)}}.demo-podium,.demo-cv-shadow{{left:44%}}
}}
@media(max-width:768px){{
    .header{{padding:1rem;gap:1rem}}
    .logo-icon{{width:60px;height:60px}}.logo-text{{font-size:1.55rem}}
    .header-middle{{grid-template-columns:1fr;padding-top:.6rem}}
    .screening-intake{{min-height:145px;padding-left:.2rem;border-right:0;border-bottom:1px solid var(--border)}}
    .outcome-grid{{min-height:175px;padding:0 .2rem;gap:.25rem}}
    .outcome-card{{min-height:165px;padding-left:1rem}}
    .outcome-card .stat-lbl{{font-size:.6rem}}
    .donut-wrap{{grid-column:auto}}
    .restrictions{{gap:.8rem 1.2rem}}.criteria-section{{min-width:0}}
    .main{{padding:1rem}}
    .criterion-lab-head{{flex-direction:column}}
    .criterion-lab-controls,.criterion-lab-results{{padding:1rem}}
    .criterion-rule{{align-items:flex-start;flex-direction:column}}
    .criterion-mode-group{{width:100%}}.criterion-mode-btn{{flex:1}}
    .criterion-result-grid{{grid-template-columns:1fr 1fr}}
    .criterion-lab-insights{{grid-template-columns:1fr}}
    .xray-shell{{padding:0}}
    .xray-dialog{{width:100vw;height:100vh;border-radius:0}}
    .xray-body{{grid-template-columns:1fr;overflow:auto}}
    .xray-source{{min-height:52vh;border-right:0;border-bottom:1px solid var(--border)}}
    .xray-evidence{{overflow:visible}}
    .xray-mark{{display:none}}
    .xray-head{{padding:.75rem}}
    .xray-candidate{{max-width:52vw}}
    .xray-nav{{padding:0 .38rem}}
    .tool-filter-head{{flex-direction:column}}
    .candidate-review{{padding:.65rem}}
    .review-hero{{align-items:flex-start;flex-direction:column}}
    .review-metrics{{width:100%}}.review-metric{{flex:1;min-width:0}}
    .decision-rationale{{grid-template-columns:1fr}}
    .review-split{{grid-template-columns:1fr}}
    .kw-item{{padding:.9rem}}
    .kw-header{{align-items:flex-start}}
    .kw-citation-head{{align-items:flex-start;flex-direction:column;gap:.2rem}}
    .kw-citation-source{{width:100%;justify-content:space-between}}
    .kw-citation .kw-relation{{text-align:left}}
    .ai-usage-overview,.ai-token-usage{{align-items:flex-start;flex-direction:column}}
    .ai-usage-overview small{{max-width:none;text-align:left}}
    .podium-column{{height:590px}}
    .demo-heading{{top:.4rem;right:1rem;left:1rem}}.demo-heading p{{display:none}}.demo-progress{{width:100px}}
    .demo-cv-float{{left:42%;top:44%;width:min(52vw,235px)}}.demo-podium,.demo-cv-shadow{{left:42%}}.demo-podium{{width:min(72vw,390px)}}
    .demo-outcomes{{top:29%;right:.45rem;width:128px;gap:7px}}.demo-outcome{{grid-template-columns:27px 1fr;gap:6px;padding:7px;border-radius:10px}}.demo-outcome-icon{{width:26px;height:26px;font-size:.75rem}}.demo-outcome strong{{font-size:.65rem}}
}}
@media(prefers-reduced-motion:reduce){{.exp-particle{{display:none}}.logo-icon,.logo-text,.outcome-bar,.donut-segment,.outcome-distribution i,.demo-cv,.demo-verdict-orb::after,.xray-excerpt.is-updated{{animation-duration:.001ms!important;animation-delay:0ms!important}}}}
/* Ambiguity-only AI verdict */
.semantic-review-container{{padding:0}}
.ai-scope{{display:flex;flex-wrap:wrap;gap:.4rem;margin-bottom:.75rem}}
.ai-scope span{{padding:.25rem .55rem;border:1px solid var(--border);border-radius:999px;background:var(--bg3);color:var(--text2);font-size:.68rem;font-weight:600}}
.semantic-review-btn{{background:linear-gradient(135deg,#d97706,var(--amber));color:#111827;border:none;padding:.7rem 1.2rem;border-radius:var(--radius-sm);cursor:pointer;font-size:.88rem;font-weight:700;transition:all .2s}}
.semantic-review-btn:hover{{transform:translateY(-1px);box-shadow:0 4px 16px rgba(245,158,11,.25)}}
.semantic-review-btn:disabled{{opacity:.5;cursor:not-allowed;transform:none}}
.semantic-review-output{{margin-top:1rem;display:flex;flex-direction:column;gap:.7rem}}
.ai-verdict{{display:flex;align-items:center;gap:.55rem;padding:.8rem;border-radius:var(--radius-sm);font-weight:750;font-size:1rem}}
.ai-verdict.pass{{background:var(--green-g);color:var(--green);border-left:3px solid var(--green)}}
.ai-verdict.fail{{background:var(--red-g);color:var(--red);border-left:3px solid var(--red)}}
.semantic-summary{{padding:.7rem .8rem;background:var(--amber-g);border-left:3px solid var(--amber);border-radius:var(--radius-sm);font-size:.86rem;line-height:1.55}}
.semantic-finding{{padding:.8rem;background:var(--bg2);border:1px solid var(--border);border-radius:var(--radius-sm)}}
.semantic-finding-head{{display:flex;justify-content:space-between;gap:.6rem;align-items:flex-start}}
.semantic-finding-title{{font-weight:650;font-size:.88rem}}
.semantic-classification{{font-size:.68rem;font-weight:700;padding:.15rem .45rem;border-radius:8px;background:var(--accent-g);color:var(--accent2);white-space:nowrap}}
.semantic-explanation{{font-size:.82rem;color:var(--text2);line-height:1.5;margin:.55rem 0}}
.semantic-quote{{font-family:var(--mono);font-size:.76rem;line-height:1.5;padding:.55rem;background:var(--bg);border-radius:6px;white-space:pre-wrap}}
.semantic-meta{{font-size:.7rem;color:var(--text3);margin-top:.45rem}}
.semantic-actions{{display:flex;align-items:center;gap:.45rem;margin-top:.65rem}}
.semantic-decision{{border:1px solid var(--border2);background:var(--bg4);color:var(--text);border-radius:6px;padding:.5rem .75rem;cursor:pointer;font-size:.78rem;font-weight:650}}
.semantic-decision.pass:hover{{border-color:var(--green);color:var(--green)}}.semantic-decision.fail:hover{{border-color:var(--red);color:var(--red)}}
.semantic-decision:disabled{{opacity:.5;cursor:not-allowed}}
.semantic-decision-status{{font-size:.72rem;font-weight:650;color:var(--text2)}}
.ai-loading{{display:flex;align-items:center;gap:.6rem;color:var(--text2);padding:.5rem 0}}
.ai-loading .spinner{{width:18px;height:18px;border:2px solid var(--border);border-top-color:var(--accent);border-radius:50%;animation:spin 1s linear infinite}}
@keyframes spin{{to{{transform:rotate(360deg)}}}}
.ai-error{{display:flex;flex-direction:column;gap:.22rem;padding:.7rem .8rem;background:var(--red-g);border:1px solid rgba(244,63,94,.22);border-radius:8px}}
.ai-error-title{{color:var(--red);font-size:.8rem}}
.ai-error-detail{{color:var(--text2);font-size:.74rem;line-height:1.45}}
.ai-usage-overview{{display:flex;align-items:center;justify-content:space-between;gap:1rem;margin-bottom:1rem;padding:.85rem 1rem;border:1px solid var(--border);border-radius:10px;background:var(--bg2)}}
.ai-usage-overview-copy{{display:flex;flex-direction:column;gap:.18rem}}
.ai-usage-overview-copy span{{font-size:.68rem;font-weight:750;text-transform:uppercase;letter-spacing:.08em;color:var(--text3)}}
.ai-usage-overview-copy strong{{font-size:1.05rem;color:var(--text)}}
.ai-usage-overview small{{max-width:55%;color:var(--text3);font-size:.72rem;text-align:right;line-height:1.45}}
.ai-token-usage{{display:flex;align-items:flex-start;justify-content:space-between;gap:1rem;padding:.75rem .85rem;border:1px solid var(--border);border-radius:8px;background:var(--bg3)}}
.ai-token-usage strong{{display:block;font-family:var(--mono);font-size:.9rem;color:var(--accent2)}}
.ai-token-usage span{{display:block;margin-top:.22rem;color:var(--text3);font-size:.7rem;line-height:1.45}}
.ai-token-usage-label{{font-size:.66rem;font-weight:750;text-transform:uppercase;letter-spacing:.08em;color:var(--text3);white-space:nowrap}}
.ai-token-unavailable{{border-style:dashed}}

/* Evidence-backed Interview Architect */
.interview-architect-section{{position:relative;overflow:hidden}}
.interview-architect-section::before{{position:absolute;top:0;right:0;width:180px;height:180px;background:radial-gradient(circle,rgba(124,92,252,.13),transparent 68%);content:"";pointer-events:none}}
.interview-architect-container{{position:relative}}
.interview-scope{{display:flex;flex-wrap:wrap;gap:.4rem;margin-bottom:.75rem}}
.interview-scope span{{padding:.25rem .55rem;border:1px solid color-mix(in srgb,var(--accent) 26%,var(--border));border-radius:999px;background:var(--accent-g);color:var(--accent2);font-size:.68rem;font-weight:650}}
.interview-architect-btn{{padding:.7rem 1.2rem;border:0;border-radius:var(--radius-sm);background:linear-gradient(135deg,var(--accent),#9b7cff);color:#fff;font:inherit;font-size:.88rem;font-weight:720;cursor:pointer;transition:transform .2s,box-shadow .2s,opacity .2s}}
.interview-architect-btn:hover{{transform:translateY(-1px);box-shadow:0 5px 18px rgba(124,92,252,.3)}}
.interview-architect-btn:disabled{{opacity:.5;cursor:not-allowed;transform:none}}
.interview-architect-output{{display:grid;gap:.8rem;margin-top:1rem}}
.interview-plan-summary{{padding:.72rem .8rem;border:1px solid color-mix(in srgb,var(--accent) 30%,var(--border));border-left:3px solid var(--accent);border-radius:8px;background:var(--accent-g);color:var(--text2);font-size:.82rem;line-height:1.5}}
.interview-groups{{display:grid;grid-template-columns:1fr;gap:.7rem;align-items:start}}
.interview-group{{min-width:0;padding:.75rem;border:1px solid var(--border);border-radius:10px;background:var(--bg3)}}
.interview-group-head{{display:flex;align-items:flex-start;justify-content:space-between;gap:.55rem;padding-bottom:.6rem;border-bottom:1px solid var(--border)}}
.interview-group-title{{font-size:.76rem;font-weight:760;line-height:1.25}}
.interview-group-copy{{display:block;margin-top:.18rem;color:var(--text3);font-size:.61rem;font-weight:500;line-height:1.35}}
.interview-group-count{{min-width:22px;padding:.12rem .38rem;border-radius:999px;background:var(--bg);color:var(--accent2);font-family:var(--mono);font-size:.62rem;text-align:center}}
.interview-question-list{{display:grid;gap:.8rem;margin-top:.75rem}}
.interview-question{{padding:1rem;border:1px solid var(--border);border-left:3px solid var(--accent);border-radius:8px;background:var(--bg2)}}
.interview-question[data-priority="HIGH"]{{border-left-color:var(--amber)}}
.interview-question-top{{display:flex;align-items:center;justify-content:space-between;gap:.5rem;margin-bottom:.45rem}}
.interview-question-number{{color:var(--text3);font-family:var(--mono);font-size:.6rem;font-weight:700}}
.interview-priority{{padding:.13rem .38rem;border-radius:999px;background:var(--accent-g);color:var(--accent2);font-size:.56rem;font-weight:780;letter-spacing:.06em}}
.interview-question[data-priority="HIGH"] .interview-priority{{background:var(--amber-g);color:var(--amber)}}
.interview-question-text{{color:var(--text);font-size:.79rem;font-weight:700;line-height:1.6}}
.interview-rationale{{margin-top:.65rem;color:var(--text2);font-size:.68rem;line-height:1.58}}
.interview-listen{{margin-top:.9rem;padding:.68rem .75rem;border-radius:7px;background:var(--green-g);color:var(--text2);font-size:.68rem;line-height:1.58}}
.interview-listen strong{{color:var(--green);font-weight:720}}
.interview-source-label{{display:block;margin-top:1rem;margin-bottom:.3rem;color:var(--text3);font-size:.57rem;font-weight:740;letter-spacing:.07em;text-transform:uppercase}}
.interview-question-top+.interview-source-label{{margin-top:.15rem}}
.interview-source{{margin-top:.28rem;padding:.48rem .55rem;border-radius:6px;background:var(--bg);color:var(--text3);font-family:var(--mono);font-size:.61rem;line-height:1.45;white-space:pre-wrap}}
.interview-no-signal{{padding:.75rem .2rem .15rem;color:var(--text3);font-size:.65rem;font-style:italic;line-height:1.45}}
.interview-usage{{color:var(--text3);font-size:.62rem;text-align:right}}

@media(max-width:768px){{.ai-usage-overview,.ai-token-usage{{align-items:flex-start;flex-direction:column}}.ai-usage-overview small{{max-width:none;text-align:left}}.interview-usage{{text-align:left}}}}
</style>
</head>
<body>
<header class="header">
    <div class="header-top">
        <div class="header-brand">
            <div class="logo-icon"><img src="icon.png" alt="Kriterion"></div>
            <div class="logo-text">Kriterion</div>
        </div>
        <div class="theme-btn" onclick="toggleTheme()" id="themeBtn">\U0001f319 Dark</div>
    </div>

    <div class="header-middle">
        <button type="button" class="stat-box stat-all stat-active screening-intake" data-f="all" aria-pressed="true" aria-label="Show all {total} candidates">
            <span class="summary-kicker">Screened intake</span>
            <span class="stat-val">{total}</span>
            <span class="stat-lbl">candidate CVs processed</span>
        </button>
        <div class="outcome-grid">
            <button type="button" class="stat-box stat-pass outcome-card" data-f="pass" aria-pressed="false" aria-label="Show {passed} passed candidates">
                <i class="outcome-bar" aria-hidden="true"></i><span class="stat-lbl">Passed</span><span class="stat-val">{passed}</span><span class="stat-pct">{pass_pct:.1f}%</span>
            </button>
            <button type="button" class="stat-box stat-fail outcome-card" data-f="fail" aria-pressed="false" aria-label="Show {failed} failed candidates">
                <i class="outcome-bar" aria-hidden="true"></i><span class="stat-lbl">Failed</span><span class="stat-val">{failed}</span><span class="stat-pct">{fail_pct:.1f}%</span>
            </button>
            <button type="button" class="stat-box stat-amb outcome-card" data-f="ambiguous" aria-pressed="false" aria-label="Show {ambiguous} ambiguous candidates">
                <i class="outcome-bar" aria-hidden="true"></i><span class="stat-lbl">Ambiguous</span><span class="stat-val">{ambiguous}</span><span class="stat-pct">{ambiguous_pct:.1f}%</span>
            </button>
        </div>
        <div class="donut-wrap">
            <div class="donut">
                <svg viewBox="0 0 36 36" aria-hidden="true">
                    <circle class="donut-track" cx="18" cy="18" r="14" pathLength="100"/>
                    <circle class="donut-segment" cx="18" cy="18" r="14" pathLength="100" stroke="var(--green)" style="--dash:{pass_pct:.2f} {100 - pass_pct:.2f};--offset:0"/>
                    <circle class="donut-segment" cx="18" cy="18" r="14" pathLength="100" stroke="var(--red)" style="--dash:{fail_pct:.2f} {100 - fail_pct:.2f};--offset:-{pass_pct:.2f}"/>
                    <circle class="donut-segment" cx="18" cy="18" r="14" pathLength="100" stroke="var(--amber)" style="--dash:{ambiguous_pct:.2f} {100 - ambiguous_pct:.2f};--offset:-{pass_pct + fail_pct:.2f}"/>
                </svg>
                <div class="donut-center"><strong>{pass_rate}%</strong><span>Pass rate</span></div>
            </div>
            <div class="donut-legend">
                <div class="legend-item" style="--dot:var(--green)"><span class="legend-dot"></span><span>Pass</span><b>{passed}</b></div>
                <div class="legend-item" style="--dot:var(--red)"><span class="legend-dot"></span><span>Fail</span><b>{failed}</b></div>
                <div class="legend-item" style="--dot:var(--amber)"><span class="legend-dot"></span><span>Ambiguous</span><b>{ambiguous}</b></div>
            </div>
        </div>
        <div class="outcome-distribution" aria-hidden="true"><i class="dist-pass" style="flex:{passed}"></i><i class="dist-fail" style="flex:{failed}"></i><i class="dist-amb" style="flex:{ambiguous}"></i></div>
    </div>

    <div class="header-bottom">
        <div class="restrictions">{restriction_groups_html}</div>
    </div>
</header>

<main class="main">
    {tool_filter_panel}
    {ai_usage_overview}
    <div class="table-wrap">
        <table>
            <thead><tr><th>Candidate</th><th>Result</th><th>Score</th><th>Exp</th><th>Keywords</th></tr></thead>
            <tbody id="tableBody">{table_rows}</tbody>
        </table>
    </div>
    <section class="criterion-lab" id="criterionLab" aria-labelledby="criterionLabTitle">
        <header class="criterion-lab-head">
            <div>
                <span class="criterion-lab-kicker">Profile simulator</span>
                <h2 id="criterionLabTitle">Kriterion Lab</h2>
                <p>Stress-test the screening profile against this cohort. Apply recalculated verdicts to this rendered report; the profile YAML, CV files, verified evidence, and human decisions remain unchanged.</p>
            </div>
            <span class="criterion-lab-safety">Report only</span>
        </header>
        <div class="criterion-lab-grid">
            <div class="criterion-lab-controls">
                <div class="criterion-panel-title"><h3>Experimental profile</h3><span>Required affects verdict · Preferred tracks gaps</span></div>
                <div class="criterion-years">
                    <div class="criterion-years-copy"><strong>Minimum relevant experience</strong><span>Overlap-safe years from parsed career history</span></div>
                    <output class="criterion-years-output" id="criterionYearsOutput" for="criterionYears">{config.MIN_DEVOPS_YEARS:g} yr</output>
                    <input id="criterionYears" type="range" min="0" max="{criterion_lab_max_years}" step="0.5" value="{config.MIN_DEVOPS_YEARS:g}" aria-label="Minimum relevant experience in years">
                </div>
                <div class="criterion-rules">{criterion_lab_rule_controls}</div>
                <div class="criterion-lab-actions">
                    <button type="button" class="criterion-reset" id="criterionReset">Reset baseline</button>
                    <button type="button" class="criterion-apply" id="criterionApply" disabled>Apply to CV results</button>
                    <span class="criterion-adjustment-count" id="criterionAdjustmentCount">Baseline profile</span>
                    <span class="criterion-apply-status" id="criterionApplyStatus" role="status">Preview only · profile YAML unchanged</span>
                </div>
            </div>
            <div class="criterion-lab-results" aria-live="polite">
                <div class="criterion-panel-title"><h3>Simulated cohort</h3><span id="criterionDeltaSummary">Matches the baseline profile</span></div>
                <div class="criterion-result-grid">
                    <div class="criterion-result criterion-result-pass"><span>Pass</span><strong id="criterionPass">0</strong><small id="criterionPassDelta"></small></div>
                    <div class="criterion-result criterion-result-fail"><span>Fail</span><strong id="criterionFail">0</strong><small id="criterionFailDelta"></small></div>
                    <div class="criterion-result criterion-result-amb"><span>Ambiguous</span><strong id="criterionAmbiguous">0</strong><small id="criterionAmbiguousDelta"></small></div>
                    <div class="criterion-result criterion-result-pref"><span>Preference gaps</span><strong id="criterionPreferenceGaps">0</strong><small>candidates</small></div>
                </div>
                <div class="criterion-distribution" aria-hidden="true"><i class="criterion-distribution-pass" id="criterionPassBar"></i><i class="criterion-distribution-fail" id="criterionFailBar"></i><i class="criterion-distribution-amb" id="criterionAmbiguousBar"></i></div>
                <div class="criterion-lab-insights">
                    <section class="criterion-subpanel">
                        <h3>Rule impact</h3>
                        <div class="criterion-impact-list">{criterion_lab_impact_rows}</div>
                    </section>
                    <div>
                        <section class="criterion-subpanel">
                            <h3>Profile signals</h3>
                            <div class="criterion-warning-list" id="criterionWarnings"></div>
                        </section>
                        <section class="criterion-subpanel" style="margin-top:.75rem">
                            <h3>Candidate movement</h3>
                            <div class="criterion-movement-list" id="criterionMovements"></div>
                        </section>
                    </div>
                </div>
            </div>
        </div>
        <footer class="criterion-lab-foot">Kriterion Lab uses the report’s verified experience evidence. Apply updates this rendered report only; the profile YAML, CV files, verified evidence, and human decisions remain unchanged.</footer>
    </section>
</main>

<section class="insights-stage" id="screeningStage" aria-label="Screening insights and scroll demonstration">
    <div class="insights-sticky">
        <div class="charts-column">
            <div class="chart-card chart-reasons">
                <h3>Top Failure Reasons</h3>
                {reasons_bars_html}
            </div>
            <div class="chart-card chart-coverage">
                <h3>Keyword Coverage</h3>
                {kw_bars_html}
            </div>
        </div>
        <div class="podium-column" id="screenDemo">
            <div class="demo-heading">
                <div><span class="demo-eyebrow">AI-assisted review</span><h2>AI helps Kriterion resolve ambiguous evidence</h2></div>
                <div class="demo-progress" aria-label="Screening animation progress"><div class="demo-progress-copy"><span>Scan</span><span id="demoProgressPercent">0%</span></div><div class="demo-progress-track"><span class="demo-progress-bar" id="demoProgressBar"></span></div></div>
            </div>
            <div class="demo-scene">
                <div class="demo-cv-shadow"></div>
                <div class="demo-cv-float">
                    <article class="demo-cv" aria-label="Abstract candidate CV being screened">
                        <header class="demo-cv-header"><div class="demo-cv-mark">CV</div><div><strong>Candidate evidence</strong><span>{profile_role}</span></div></header>
                        <section class="demo-cv-section"><h4>Experience</h4><div class="demo-role-row"><div><div class="demo-line demo-line-long"></div><div class="demo-line demo-line-mid"></div></div></div><div class="demo-role-row"><div><div class="demo-line demo-line-mid"></div><div class="demo-line demo-line-short"></div></div></div></section>
                        <section class="demo-cv-section"><h4>Required keyword coverage</h4>{podium_skills_html}</section>
                        <section class="demo-cv-section"><h4>Career history</h4><div class="demo-line demo-line-long"></div><div class="demo-line demo-line-mid"></div><div class="demo-line demo-line-short"></div></section>
                        <div class="demo-scan-beam" aria-hidden="true"><div class="demo-scan-line"></div><div class="demo-scan-wash"></div></div>
                    </article>
                    <div class="demo-verdict-orb" aria-hidden="true"></div>
                </div>
                <div class="demo-podium" aria-hidden="true"><div class="demo-podium-top"><div class="demo-podium-core"></div></div><div class="demo-podium-rim"></div><div class="demo-podium-base"></div><div class="demo-podium-reflection"></div></div>
                <div class="demo-outcomes" aria-label="Real report outcomes">
                    <article class="demo-outcome demo-outcome-pass"><div class="demo-outcome-icon">✓</div><div><span>Passed</span><strong>{passed}</strong></div></article>
                    <article class="demo-outcome demo-outcome-fail"><div class="demo-outcome-icon">×</div><div><span>Failed</span><strong>{failed}</strong></div></article>
                    <article class="demo-outcome demo-outcome-amb"><div class="demo-outcome-icon">?</div><div><span>Ambiguous</span><strong>{ambiguous}</strong></div></article>
                </div>
            </div>
            <p class="demo-caption">Profile-driven animation · aggregate values are from this report</p>
        </div>
    </div>
</section>

<div class="xray-shell" id="evidenceXray" hidden aria-hidden="true">
    <button type="button" class="xray-backdrop" id="xrayBackdrop" aria-label="Close Evidence X-Ray"></button>
    <section class="xray-dialog" role="dialog" aria-modal="true" aria-labelledby="xrayTitle" tabindex="-1">
        <header class="xray-head">
            <div class="xray-title-group">
                <div class="xray-mark" aria-hidden="true">⌖</div>
                <div><span class="xray-kicker">Evidence X-Ray</span><h2 id="xrayTitle">Verified requirement evidence</h2><span class="xray-candidate" id="xrayCandidate"></span></div>
            </div>
            <div class="xray-head-actions">
                <button type="button" class="xray-nav" id="xrayPrevious" aria-label="Previous evidence">← Prev</button>
                <button type="button" class="xray-nav" id="xrayNext" aria-label="Next evidence">Next →</button>
                <button type="button" class="xray-close" id="xrayClose" aria-label="Close Evidence X-Ray">×</button>
            </div>
        </header>
        <div class="xray-body">
            <section class="xray-source" aria-label="Original CV source">
                <div class="xray-source-head"><span>Original CV</span><span class="xray-page" id="xrayPageLabel"></span></div>
                <div class="xray-preview" id="xrayPreview" hidden><span class="xray-loading" id="xrayLoading">Rendering cited page and verified highlight…</span><img class="xray-page-image" id="xrayPageImage" alt=""></div>
                <div class="xray-source-fallback" id="xrayFallback" hidden><strong id="xrayFallbackTitle">Source preview unavailable</strong><span id="xrayFallbackCopy"></span><a class="xray-open-original" id="xrayFallbackLink" target="_blank" rel="noopener">Open original CV</a></div>
            </section>
            <aside class="xray-evidence" aria-label="Verified evidence details">
                <div class="xray-badges"><span class="xray-badge xray-badge-primary" id="xrayRequirement"></span><span class="xray-badge" id="xrayRelationship"></span><span class="xray-badge" id="xrayMatchedTerm"></span></div>
                <div class="xray-facts">
                    <div class="xray-fact"><span>Candidate</span><strong id="xrayFactCandidate"></strong></div>
                    <div class="xray-fact"><span>Source page</span><strong id="xrayFactPage"></strong></div>
                    <div class="xray-fact"><span>Role context</span><strong id="xrayRoleContext"></strong></div>
                </div>
                <h3>Exact extracted evidence</h3>
                <blockquote class="xray-excerpt" id="xrayExcerpt"></blockquote>
                <div class="xray-note">The matched term is highlighted in the verified extracted excerpt. The source viewer opens the cited PDF page for direct comparison.</div>
                <a class="xray-open-original" id="xrayOpenOriginal" target="_blank" rel="noopener" style="margin-top:.75rem">Open cited page in a new tab</a>
            </aside>
        </div>
    </section>
</div>

<script type="application/json" id="criterionLabData">{criterion_lab_data_json}</script>
<script type="application/json" id="evidenceXrayData">{evidence_xray_data_json}</script>
<script>
// Theme
function toggleTheme(){{
    const h=document.documentElement,t=h.dataset.theme==='dark'?'light':'dark';
    h.dataset.theme=t;localStorage.setItem('rt-theme',t);
    document.getElementById('themeBtn').textContent=t==='dark'?'\U0001f319 Dark':'☀️ Light';
}}
(()=>{{const s=localStorage.getItem('rt-theme');if(s){{document.documentElement.dataset.theme=s;document.getElementById('themeBtn').textContent=s==='dark'?'\U0001f319 Dark':'☀️ Light';}}}})();

// Evidence X-Ray: exact excerpt plus the original CV opened at its cited page.
var evidenceXrayDataNode=document.getElementById('evidenceXrayData');
var evidenceXrayItems=evidenceXrayDataNode?JSON.parse(evidenceXrayDataNode.textContent):[];
var evidenceXrayIndexById=new Map(evidenceXrayItems.map(function(item,index){{return[item.id,index];}}));
var evidenceXrayIndicesByCandidate=new Map();
evidenceXrayItems.forEach(function(item,index){{
    if(!evidenceXrayIndicesByCandidate.has(item.candidate))evidenceXrayIndicesByCandidate.set(item.candidate,[]);
    evidenceXrayIndicesByCandidate.get(item.candidate).push(index);
}});
var evidenceXrayCurrentIndex=-1;
var evidenceXrayPreviousFocus=null;
var evidenceXrayFrameToken=0;

function xraySetText(id,value){{var node=document.getElementById(id);if(node)node.textContent=String(value||'');}}
function xraySourceUrl(item){{
    return item.sourceUrl?item.sourceUrl+'#page='+encodeURIComponent(item.page)+'&zoom=page-width&search='+encodeURIComponent(item.matchedTerm||item.requirement||''):'';
}}
function xrayPreviewUrl(item){{
    return item.previewUrl?item.previewUrl+'?page='+encodeURIComponent(item.page)+'&term='+encodeURIComponent(item.matchedTerm||item.requirement||''):'';
}}
function renderXrayExcerpt(text,matchedTerm){{
    var container=document.getElementById('xrayExcerpt');
    if(!container)return 0;
    container.classList.remove('is-updated');
    container.innerHTML='';
    var source=String(text||''),term=String(matchedTerm||'');
    if(!term){{container.appendChild(document.createTextNode(source));void container.offsetWidth;container.classList.add('is-updated');return 0;}}
    var lowerSource=source.toLowerCase(),lowerTerm=term.toLowerCase(),cursor=0,index=lowerSource.indexOf(lowerTerm),matches=0;
    while(index!==-1){{
        if(index>cursor)container.appendChild(document.createTextNode(source.slice(cursor,index)));
        var mark=document.createElement('mark');mark.textContent=source.slice(index,index+term.length);container.appendChild(mark);
        matches+=1;cursor=index+term.length;index=lowerSource.indexOf(lowerTerm,cursor);
    }}
    if(cursor<source.length)container.appendChild(document.createTextNode(source.slice(cursor)));
    void container.offsetWidth;container.classList.add('is-updated');
    return matches;
}}
function xrayCandidateSequence(index){{
    var item=evidenceXrayItems[index];
    return item?(evidenceXrayIndicesByCandidate.get(item.candidate)||[]):[];
}}
function navigateEvidenceXray(offset){{
    var sequence=xrayCandidateSequence(evidenceXrayCurrentIndex);
    var position=sequence.indexOf(evidenceXrayCurrentIndex);
    var nextPosition=position+offset;
    if(position===-1||nextPosition<0||nextPosition>=sequence.length)return;
    renderEvidenceXray(sequence[nextPosition]);
}}
function loadXrayPagePreview(image,loading,previewUrl,item){{
    evidenceXrayFrameToken+=1;
    var token=evidenceXrayFrameToken;
    image.classList.add('is-loading');
    image.alt='Original CV page '+item.page+' for '+item.candidate+' with '+item.matchedTerm+' highlighted';
    image.onload=function(){{if(token!==evidenceXrayFrameToken)return;image.classList.remove('is-loading');if(loading)loading.hidden=true;}};
    image.onerror=function(){{
        if(token!==evidenceXrayFrameToken)return;
        var preview=document.getElementById('xrayPreview'),fallback=document.getElementById('xrayFallback');
        if(preview)preview.hidden=true;if(fallback)fallback.hidden=false;
        xraySetText('xrayFallbackTitle','Highlighted preview unavailable');
        xraySetText('xrayFallbackCopy','Kriterion could not render this cited PDF page. Open the original document and compare it with the highlighted verified excerpt.');
    }};
    image.removeAttribute('src');
    if(loading)loading.hidden=false;
    requestAnimationFrame(function(){{
        if(token===evidenceXrayFrameToken)image.src=previewUrl;
    }});
}}
function renderEvidenceXray(index){{
    if(index<0||index>=evidenceXrayItems.length)return;
    evidenceXrayCurrentIndex=index;
    var item=evidenceXrayItems[index],sourceUrl=xraySourceUrl(item),previewUrl=xrayPreviewUrl(item);
    xraySetText('xrayCandidate',item.candidate);
    xraySetText('xrayRequirement',item.requirement);
    xraySetText('xrayRelationship',item.relationship||'direct evidence');
    xraySetText('xrayFactCandidate',item.candidate);
    xraySetText('xrayFactPage','Page '+item.page);
    xraySetText('xrayRoleContext',item.roleContext||'Role header not confidently mapped');
    var highlightedMatches=renderXrayExcerpt(item.snippet,item.matchedTerm);
    xraySetText('xrayMatchedTerm','Matched: '+item.matchedTerm+(highlightedMatches?' · highlighted':' · verify manually'));
    var sequence=xrayCandidateSequence(index),position=sequence.indexOf(index);
    xraySetText('xrayPageLabel','Cited page '+item.page+' · Evidence '+(position+1)+' of '+sequence.length);
    var preview=document.getElementById('xrayPreview');
    var image=document.getElementById('xrayPageImage');
    var loading=document.getElementById('xrayLoading');
    var fallback=document.getElementById('xrayFallback');
    var fallbackLink=document.getElementById('xrayFallbackLink');
    var openOriginal=document.getElementById('xrayOpenOriginal');
    if(preview)preview.hidden=!item.previewAvailable;
    if(image){{
        if(item.previewAvailable)loadXrayPagePreview(image,loading,previewUrl,item);
        else{{evidenceXrayFrameToken+=1;image.removeAttribute('src');}}
    }}
    if(fallback)fallback.hidden=Boolean(item.previewAvailable);
    if(!item.previewAvailable){{
        xraySetText('xrayFallbackTitle',item.sourceUrl?'Document preview unavailable':'Original CV link unavailable');
        xraySetText('xrayFallbackCopy',item.sourceUrl?'This source format cannot be embedded. Open the original document and compare it with the verified excerpt.':'This static report was generated without a served CV folder. The verified extracted excerpt remains available for review.');
    }}
    [fallbackLink,openOriginal].forEach(function(link){{
        if(!link)return;
        link.hidden=!item.sourceUrl;
        if(item.sourceUrl)link.href=sourceUrl;else link.removeAttribute('href');
    }});
    var previous=document.getElementById('xrayPrevious'),next=document.getElementById('xrayNext');
    if(previous)previous.disabled=position<=0;
    if(next)next.disabled=position<0||position>=sequence.length-1;
}}
function openEvidenceXray(id,trigger){{
    var index=evidenceXrayIndexById.get(id),shell=document.getElementById('evidenceXray');
    if(index===undefined||!shell)return;
    evidenceXrayPreviousFocus=trigger||document.activeElement;
    renderEvidenceXray(index);
    shell.hidden=false;shell.setAttribute('aria-hidden','false');document.body.classList.add('xray-open');
    var close=document.getElementById('xrayClose');if(close)close.focus();
}}
function closeEvidenceXray(){{
    var shell=document.getElementById('evidenceXray'),image=document.getElementById('xrayPageImage');
    if(!shell||shell.hidden)return;
    shell.hidden=true;shell.setAttribute('aria-hidden','true');document.body.classList.remove('xray-open');
    evidenceXrayFrameToken+=1;if(image)image.removeAttribute('src');
    if(evidenceXrayPreviousFocus&&typeof evidenceXrayPreviousFocus.focus==='function')evidenceXrayPreviousFocus.focus();
}}
document.querySelectorAll('.xray-open-btn').forEach(function(button){{
    button.addEventListener('click',function(event){{event.preventDefault();event.stopPropagation();openEvidenceXray(button.dataset.xrayId,button);}});
}});
var xrayClose=document.getElementById('xrayClose'),xrayBackdrop=document.getElementById('xrayBackdrop');
if(xrayClose)xrayClose.addEventListener('click',closeEvidenceXray);
if(xrayBackdrop)xrayBackdrop.addEventListener('click',closeEvidenceXray);
var xrayPrevious=document.getElementById('xrayPrevious'),xrayNext=document.getElementById('xrayNext');
if(xrayPrevious)xrayPrevious.addEventListener('click',function(){{navigateEvidenceXray(-1);}});
if(xrayNext)xrayNext.addEventListener('click',function(){{navigateEvidenceXray(1);}});
document.addEventListener('keydown',function(event){{
    var shell=document.getElementById('evidenceXray');
    if(!shell||shell.hidden)return;
    if(event.key==='Escape'){{event.preventDefault();closeEvidenceXray();return;}}
    if(event.key==='ArrowLeft'){{event.preventDefault();navigateEvidenceXray(-1);}}
    if(event.key==='ArrowRight'){{event.preventDefault();navigateEvidenceXray(1);}}
    if(event.key==='Tab'){{
        var focusable=Array.from(shell.querySelectorAll('button:not(:disabled),a[href]')).filter(function(node){{return node.offsetParent!==null;}});
        if(!focusable.length)return;
        var first=focusable[0],last=focusable[focusable.length-1];
        if(event.shiftKey&&document.activeElement===first){{event.preventDefault();last.focus();}}
        else if(!event.shiftKey&&document.activeElement===last){{event.preventDefault();first.focus();}}
    }}
}});

// Native scroll-driven podium screening demonstration
const screeningStage=document.getElementById('screeningStage');
const screeningDemo=document.getElementById('screenDemo');
const demoProgressBar=document.getElementById('demoProgressBar');
const demoProgressPercent=document.getElementById('demoProgressPercent');
const demoPassTarget=screeningDemo?screeningDemo.querySelector('.demo-outcome-pass .demo-outcome-icon'):null;
const demoFailTarget=screeningDemo?screeningDemo.querySelector('.demo-outcome-fail .demo-outcome-icon'):null;
const demoAmbTarget=screeningDemo?screeningDemo.querySelector('.demo-outcome-amb .demo-outcome-icon'):null;
let screeningTicking=false;
const screeningClamp=(value,min=0,max=1)=>Math.min(max,Math.max(min,value));
const screeningRange=(value,start,end)=>screeningClamp((value-start)/(end-start));
const screeningEase=value=>value*value*(3-2*value);
const screeningMix=(from,to,amount)=>from+(to-from)*amount;
function screeningTargetOffset(target){{
    if(!target||!screeningDemo)return{{x:0,y:0}};
    const demoRect=screeningDemo.getBoundingClientRect();
    const targetRect=target.getBoundingClientRect();
    return{{
        x:targetRect.left+targetRect.width/2-(demoRect.left+screeningDemo.clientWidth*.42),
        y:targetRect.top+targetRect.height/2-(demoRect.top+screeningDemo.clientHeight*.43)
    }};
}}
function renderScreeningDemo(){{
    screeningTicking=false;
    if(!screeningStage||!screeningDemo)return;
    const compact=window.matchMedia('(max-width:900px)').matches;
    const reduced=window.matchMedia('(prefers-reduced-motion:reduce)').matches;
    const scrollable=screeningStage.offsetHeight-window.innerHeight;
    let progress=1;
    if(!compact&&!reduced&&scrollable>0){{
        progress=screeningClamp(-screeningStage.getBoundingClientRect().top/scrollable);
    }}
    const enter=screeningEase(screeningRange(progress,.03,.19));
    const reasons=screeningEase(screeningRange(progress,.10,.30));
    const scan=screeningEase(screeningRange(progress,.22,.68));
    const coverage=screeningEase(screeningRange(progress,.36,.58));
    const morph=screeningEase(screeningRange(progress,.62,.70));
    const decision=screeningEase(screeningRange(progress,.64,.72));
    const tiltY=(-10+scan*2)*(1-morph);
    const tiltX=(4-scan*1.5)*(1-morph);
    const depth=70*(1-morph);
    const passTarget=screeningTargetOffset(demoPassTarget);
    const failTarget=screeningTargetOffset(demoFailTarget);
    const ambTarget=screeningTargetOffset(demoAmbTarget);
    let flightX=0,flightY=0,orbColor='var(--accent)';
    if(progress>=.70&&progress<.82){{
        const travel=screeningEase(screeningRange(progress,.70,.77));
        flightX=screeningMix(0,passTarget.x,travel);
        flightY=screeningMix(0,passTarget.y,travel);
        orbColor='var(--green)';
    }}else if(progress>=.82&&progress<.90){{
        const travel=screeningEase(screeningRange(progress,.82,.87));
        flightX=screeningMix(passTarget.x,failTarget.x,travel);
        flightY=screeningMix(passTarget.y,failTarget.y,travel);
        orbColor='var(--red)';
    }}else if(progress>=.90){{
        const travel=screeningEase(screeningRange(progress,.90,.95));
        flightX=screeningMix(failTarget.x,ambTarget.x,travel);
        flightY=screeningMix(failTarget.y,ambTarget.y,travel);
        orbColor='var(--amber)';
    }}
    const passGlow=screeningEase(screeningRange(progress,.755,.785))*(1-screeningEase(screeningRange(progress,.81,.835)));
    const failGlow=screeningEase(screeningRange(progress,.855,.88))*(1-screeningEase(screeningRange(progress,.89,.915)));
    const ambGlow=screeningEase(screeningRange(progress,.94,.97));
    screeningStage.style.setProperty('--demo-enter',enter.toFixed(4));
    screeningStage.style.setProperty('--demo-reasons',reasons.toFixed(4));
    screeningStage.style.setProperty('--demo-scan',scan.toFixed(4));
    screeningStage.style.setProperty('--demo-coverage',coverage.toFixed(4));
    screeningStage.style.setProperty('--demo-decision',decision.toFixed(4));
    screeningStage.style.setProperty('--demo-morph',morph.toFixed(4));
    screeningStage.style.setProperty('--demo-tilt-y',tiltY.toFixed(3)+'deg');
    screeningStage.style.setProperty('--demo-tilt-x',tiltX.toFixed(3)+'deg');
    screeningStage.style.setProperty('--demo-depth',depth.toFixed(2)+'px');
    screeningStage.style.setProperty('--demo-flight-x',flightX.toFixed(2)+'px');
    screeningStage.style.setProperty('--demo-flight-y',flightY.toFixed(2)+'px');
    screeningStage.style.setProperty('--demo-pass-glow',passGlow.toFixed(4));
    screeningStage.style.setProperty('--demo-fail-glow',failGlow.toFixed(4));
    screeningStage.style.setProperty('--demo-amb-glow',ambGlow.toFixed(4));
    screeningStage.style.setProperty('--demo-orb-color',orbColor);
    if(demoProgressBar)demoProgressBar.style.transform=`scaleX(${{progress}})`;
    if(demoProgressPercent)demoProgressPercent.textContent=`${{Math.round(progress*100)}}%`;
}}
function requestScreeningRender(){{
    if(!screeningTicking){{screeningTicking=true;requestAnimationFrame(renderScreeningDemo);}}
}}
window.addEventListener('scroll',requestScreeningRender,{{passive:true}});
window.addEventListener('resize',requestScreeningRender);
renderScreeningDemo();

// Kriterion Lab: isolated cohort simulation over verified report evidence.
var criterionLabNode=document.getElementById('criterionLabData');
var criterionLabPayload=criterionLabNode?JSON.parse(criterionLabNode.textContent):{{minimumYears:0,rules:[],candidates:[]}};
var criterionLabDefault={{minimumYears:Number(criterionLabPayload.minimumYears)||0,rules:{{}}}};
criterionLabPayload.rules.forEach(function(rule){{criterionLabDefault.rules[rule]='required';}});
var criterionLabState={{minimumYears:criterionLabDefault.minimumYears,rules:Object.assign({{}},criterionLabDefault.rules)}};
function criterionLabStateKey(state){{
    return JSON.stringify({{minimumYears:Number(state.minimumYears),rules:criterionLabPayload.rules.map(function(rule){{return[rule,state.rules[rule]];}})}});
}}
var criterionLabAppliedStateKey=criterionLabStateKey(criterionLabDefault);
var criterionLabHasApplied=false;

function criterionLabClassify(candidate,state){{
    var requiredRules=criterionLabPayload.rules.filter(function(rule){{return state.rules[rule]==='required';}});
    var needsReview=Boolean(candidate.layoutReview||candidate.dateReview)||requiredRules.some(function(rule){{
        return candidate.evidence[rule]==='review';
    }});
    if(needsReview)return 'ambiguous';
    var missingRequired=requiredRules.some(function(rule){{return candidate.evidence[rule]!=='found';}});
    var fixedFailure=Array.isArray(candidate.fixedFailures)&&candidate.fixedFailures.length>0;
    if(fixedFailure||missingRequired||Number(candidate.years)<Number(state.minimumYears))return 'fail';
    return 'pass';
}}

function criterionLabEvaluate(state){{
    var counts={{pass:0,fail:0,ambiguous:0,preferenceGaps:0,statuses:[]}};
    var preferredRules=criterionLabPayload.rules.filter(function(rule){{return state.rules[rule]==='preferred';}});
    criterionLabPayload.candidates.forEach(function(candidate){{
        var status=criterionLabClassify(candidate,state);
        counts[status]+=1;
        counts.statuses.push(status);
        if(preferredRules.some(function(rule){{return candidate.evidence[rule]!=='found';}}))counts.preferenceGaps+=1;
    }});
    return counts;
}}

var criterionLabBaseline=criterionLabEvaluate(criterionLabDefault);
function criterionLabDelta(value,baseline){{
    var delta=value-baseline;
    return delta===0?'no change':(delta>0?'+':'')+delta+' vs baseline';
}}
function criterionLabSetText(id,value){{
    var node=document.getElementById(id);
    if(node)node.textContent=String(value);
}}
function criterionLabAlternative(rule,mode){{
    var state={{minimumYears:criterionLabState.minimumYears,rules:Object.assign({{}},criterionLabState.rules)}};
    if(rule==='__years__')state.minimumYears=0;else state.rules[rule]=mode;
    return criterionLabEvaluate(state);
}}
function criterionLabImpactRow(rule,current){{
    var row=document.querySelector('.criterion-impact-row[data-impact-rule="'+CSS.escape(rule)+'"]');
    if(!row)return;
    var copy=row.querySelector('.criterion-impact-copy');
    var bar=row.querySelector('.criterion-impact-track i');
    var total=Math.max(criterionLabPayload.candidates.length,1);
    var affected=0,passGain=0,message='';
    if(rule==='__years__'){{
        affected=criterionLabPayload.candidates.filter(function(candidate){{return Number(candidate.years)<criterionLabState.minimumYears;}}).length;
        passGain=Math.max(0,criterionLabAlternative(rule,'off').pass-current.pass);
        message=criterionLabState.minimumYears===0?'Inactive':' '+affected+' below · +'+passGain+' passes if removed';
    }}else{{
        var mode=criterionLabState.rules[rule];
        affected=criterionLabPayload.candidates.filter(function(candidate){{return candidate.evidence[rule]!=='found';}}).length;
        if(mode==='required'){{
            passGain=Math.max(0,criterionLabAlternative(rule,'preferred').pass-current.pass);
            message=affected+' blocked/reviewed · +'+passGain+' passes if relaxed';
        }}else if(mode==='preferred'){{
            message=affected+' preference gaps · no verdict effect';
        }}else{{
            affected=0;
            message='Not evaluated';
        }}
    }}
    if(copy)copy.textContent=message.trim();
    if(bar)bar.style.width=Math.round(affected/total*100)+'%';
}}
function criterionLabRenderWarnings(current){{
    var container=document.getElementById('criterionWarnings');
    if(!container)return;
    container.innerHTML='';
    var total=criterionLabPayload.candidates.length;
    var warnings=[];
    var requiredRules=criterionLabPayload.rules.filter(function(rule){{return criterionLabState.rules[rule]==='required';}});
    if(total===0){{
        warnings.push({{text:'No candidates are available for simulation.',positive:false}});
    }}else{{
        var passRate=current.pass/total;
        if(current.pass===0)warnings.push({{text:'Zero-pass profile: every candidate is currently blocked or unresolved.',positive:false}});
        else if(passRate<.1)warnings.push({{text:'Highly restrictive profile: fewer than 10% of candidates pass.',positive:false}});
        if(requiredRules.length===0)warnings.push({{text:'No technology gate is active; verdicts rely on experience and fixed restrictions.',positive:false}});
        var bottlenecks=requiredRules.map(function(rule){{
            return{{rule:rule,count:criterionLabPayload.candidates.filter(function(candidate){{return candidate.evidence[rule]!=='found';}}).length}};
        }}).sort(function(a,b){{return b.count-a.count;}});
        if(bottlenecks.length&&bottlenecks[0].count/total>=.7){{
            warnings.push({{text:bottlenecks[0].rule+' is the strongest bottleneck, affecting '+bottlenecks[0].count+' candidates.',positive:false}});
        }}
        var passChange=current.pass-criterionLabBaseline.pass;
        if(passChange>0)warnings.push({{text:'This experiment adds '+passChange+' potential pass'+(passChange===1?'':'es')+' versus baseline.',positive:true}});
        if(!warnings.length)warnings.push({{text:'The experimental profile has no severe cohort-level bottleneck.',positive:true}});
    }}
    warnings.slice(0,3).forEach(function(item){{
        var node=document.createElement('div');
        node.className='criterion-warning'+(item.positive?' is-positive':'');
        node.textContent=item.text;
        container.appendChild(node);
    }});
}}
function criterionLabRenderMovements(current){{
    var container=document.getElementById('criterionMovements');
    if(!container)return;
    container.innerHTML='';
    var movements=[];
    current.statuses.forEach(function(status,index){{
        var before=criterionLabBaseline.statuses[index];
        if(status!==before)movements.push({{name:criterionLabPayload.candidates[index].name,before:before,after:status}});
    }});
    if(!movements.length){{
        var empty=document.createElement('div');
        empty.className='criterion-empty';
        empty.textContent='No candidate verdicts move under this experiment.';
        container.appendChild(empty);
        return;
    }}
    movements.slice(0,5).forEach(function(item){{
        var row=document.createElement('div');row.className='criterion-movement';
        var name=document.createElement('span');name.className='criterion-movement-name';name.textContent=item.name;
        var change=document.createElement('span');change.className='criterion-movement-change';change.textContent=item.before.toUpperCase()+' → '+item.after.toUpperCase();
        change.style.color=item.after==='pass'?'var(--green)':item.after==='fail'?'var(--red)':'var(--amber)';
        row.appendChild(name);row.appendChild(change);container.appendChild(row);
    }});
    if(movements.length>5){{
        var more=document.createElement('div');more.className='criterion-movement-more';more.textContent='+'+(movements.length-5)+' more candidate changes';container.appendChild(more);
    }}
}}
function updateCriterionLab(){{
    var current=criterionLabEvaluate(criterionLabState);
    criterionLabSetText('criterionYearsOutput',Number(criterionLabState.minimumYears).toFixed(1).replace('.0','')+' yr');
    criterionLabSetText('criterionPass',current.pass);
    criterionLabSetText('criterionFail',current.fail);
    criterionLabSetText('criterionAmbiguous',current.ambiguous);
    criterionLabSetText('criterionPreferenceGaps',current.preferenceGaps);
    criterionLabSetText('criterionPassDelta',criterionLabDelta(current.pass,criterionLabBaseline.pass));
    criterionLabSetText('criterionFailDelta',criterionLabDelta(current.fail,criterionLabBaseline.fail));
    criterionLabSetText('criterionAmbiguousDelta',criterionLabDelta(current.ambiguous,criterionLabBaseline.ambiguous));
    var adjustments=(criterionLabState.minimumYears===criterionLabDefault.minimumYears?0:1)+criterionLabPayload.rules.filter(function(rule){{return criterionLabState.rules[rule]!==criterionLabDefault.rules[rule];}}).length;
    criterionLabSetText('criterionAdjustmentCount',adjustments?adjustments+' adjustment'+(adjustments===1?'':'s')+' from baseline':'Baseline profile');
    var moved=current.statuses.filter(function(status,index){{return status!==criterionLabBaseline.statuses[index];}}).length;
    criterionLabSetText('criterionDeltaSummary',moved?moved+' candidate verdict'+(moved===1?' moves':'s move'):'Matches the baseline profile');
    var passBar=document.getElementById('criterionPassBar'),failBar=document.getElementById('criterionFailBar'),ambBar=document.getElementById('criterionAmbiguousBar');
    if(passBar)passBar.style.flex=String(current.pass);
    if(failBar)failBar.style.flex=String(current.fail);
    if(ambBar)ambBar.style.flex=String(current.ambiguous);
    criterionLabImpactRow('__years__',current);
    criterionLabPayload.rules.forEach(function(rule){{criterionLabImpactRow(rule,current);}});
    criterionLabRenderWarnings(current);
    criterionLabRenderMovements(current);
    var applyButton=document.getElementById('criterionApply');
    var applyStatus=document.getElementById('criterionApplyStatus');
    var hasPendingChanges=criterionLabStateKey(criterionLabState)!==criterionLabAppliedStateKey;
    if(applyButton)applyButton.disabled=!hasPendingChanges;
    if(applyStatus){{
        applyStatus.classList.toggle('is-applied',!hasPendingChanges&&criterionLabHasApplied);
        applyStatus.textContent=hasPendingChanges?'Preview changed · Apply to update CV results; profile YAML stays unchanged':criterionLabHasApplied?'Applied to this report · profile YAML unchanged':'Preview only · profile YAML unchanged';
    }}
}}

function criterionLabUpdateDashboard(current){{
    var total=criterionLabPayload.candidates.length;
    var statusUi={{
        pass:{{count:current.pass,box:'.stat-pass',demo:'.demo-outcome-pass strong',label:'passed'}},
        fail:{{count:current.fail,box:'.stat-fail',demo:'.demo-outcome-fail strong',label:'failed'}},
        ambiguous:{{count:current.ambiguous,box:'.stat-amb',demo:'.demo-outcome-amb strong',label:'ambiguous'}}
    }};
    Object.keys(statusUi).forEach(function(status){{
        var ui=statusUi[status],pct=total?ui.count/total*100:0,box=document.querySelector(ui.box);
        if(box){{
            var value=box.querySelector('.stat-val'),percent=box.querySelector('.stat-pct');
            if(value)value.textContent=String(ui.count);
            if(percent)percent.textContent=pct.toFixed(1)+'%';
            box.setAttribute('aria-label','Show '+ui.count+' '+ui.label+' candidates');
        }}
        var demo=document.querySelector(ui.demo);
        if(demo)demo.textContent=String(ui.count);
    }});
    var passPct=total?current.pass/total*100:0,failPct=total?current.fail/total*100:0,ambiguousPct=total?current.ambiguous/total*100:0;
    var donutValue=document.querySelector('.donut-center strong');
    if(donutValue)donutValue.textContent=Math.round(passPct)+'%';
    var segments=document.querySelectorAll('.donut-segment');
    var segmentValues=[passPct,failPct,ambiguousPct],segmentOffsets=[0,-passPct,-passPct-failPct];
    segments.forEach(function(segment,index){{
        var value=segmentValues[index]||0;
        segment.style.setProperty('--dash',value.toFixed(2)+' '+(100-value).toFixed(2));
        segment.style.setProperty('--offset',String(segmentOffsets[index]||0));
    }});
    var legendValues=[current.pass,current.fail,current.ambiguous];
    document.querySelectorAll('.donut-legend .legend-item b').forEach(function(node,index){{node.textContent=String(legendValues[index]||0);}});
    var distributionValues=[current.pass,current.fail,current.ambiguous];
    document.querySelectorAll('.outcome-distribution i').forEach(function(node,index){{node.style.flex=String(distributionValues[index]||0);}});
}}

function criterionLabApplyResults(){{
    var current=criterionLabEvaluate(criterionLabState);
    var appliedExperiment=criterionLabStateKey(criterionLabState)!==criterionLabStateKey(criterionLabDefault);
    current.statuses.forEach(function(status,index){{
        var row=document.querySelector('#tableBody .data-row[data-idx="'+index+'"]');
        var detail=document.querySelector('#tableBody .detail-row[data-idx="'+index+'"]');
        var label=status.toUpperCase();
        if(row){{
            row.dataset.status=status;
            row.classList.remove('row-pass','row-fail','row-ambiguous');
            row.classList.add('row-'+status);
            var pill=row.querySelector('.pill');
            if(pill){{pill.classList.remove('pill-pass','pill-fail','pill-ambiguous');pill.classList.add('pill-'+status);pill.textContent=label;}}
            var scoreFill=row.querySelector('.score-bar-fill');
            if(scoreFill){{scoreFill.classList.remove('score-fill-pass','score-fill-fail','score-fill-ambiguous');scoreFill.classList.add('score-fill-'+status);}}
        }}
        if(detail){{
            detail.dataset.status=status;
            var hero=detail.querySelector('.review-hero');
            if(hero){{
                hero.classList.remove('review-hero-pass','review-hero-fail','review-hero-ambiguous');
                hero.classList.add('review-hero-'+status);
                var icon=hero.querySelector('.outcome-icon'),title=hero.querySelector('h3'),copy=hero.querySelector('.review-outcome p'),eyebrow=hero.querySelector('.review-eyebrow');
                if(icon)icon.textContent=status==='pass'?'✓':status==='fail'?'✗':'⚠';
                if(title)title.textContent=label;
                if(copy)copy.textContent=appliedExperiment?'Recalculated from Kriterion Lab changes; profile YAML unchanged':status==='pass'?'Meets the configured screening requirements':status==='fail'?'Does not meet one or more screening requirements':'Needs a reviewer to resolve uncertain evidence';
                if(eyebrow)eyebrow.textContent=appliedExperiment?'Kriterion Lab outcome':'Screening outcome';
            }}
            var rationale=detail.querySelector('.decision-rationale');
            if(rationale){{rationale.classList.remove('rationale-pass','rationale-fail','rationale-ambiguous');rationale.classList.add('rationale-'+status);}}
            var note=detail.querySelector('.lab-application-note');
            if(note)note.hidden=!appliedExperiment;
        }}
    }});
    criterionLabUpdateDashboard(current);
    criterionLabAppliedStateKey=criterionLabStateKey(criterionLabState);
    criterionLabHasApplied=true;
    applyCandidateFilters();
    updateCriterionLab();
}}

var criterionYears=document.getElementById('criterionYears');
if(criterionYears)criterionYears.addEventListener('input',function(){{criterionLabState.minimumYears=Number(criterionYears.value);updateCriterionLab();}});
document.querySelectorAll('.criterion-mode-btn').forEach(function(button){{
    button.addEventListener('click',function(){{
        var row=button.closest('[data-criterion-rule]');
        if(!row)return;
        var rule=row.dataset.criterionRule;
        criterionLabState.rules[rule]=button.dataset.mode;
        row.querySelectorAll('.criterion-mode-btn').forEach(function(peer){{var active=peer===button;peer.classList.toggle('is-active',active);peer.setAttribute('aria-pressed',active?'true':'false');}});
        updateCriterionLab();
    }});
}});
var criterionReset=document.getElementById('criterionReset');
if(criterionReset)criterionReset.addEventListener('click',function(){{
    criterionLabState={{minimumYears:criterionLabDefault.minimumYears,rules:Object.assign({{}},criterionLabDefault.rules)}};
    if(criterionYears)criterionYears.value=String(criterionLabState.minimumYears);
    document.querySelectorAll('[data-criterion-rule]').forEach(function(row){{
        row.querySelectorAll('.criterion-mode-btn').forEach(function(button){{var active=button.dataset.mode==='required';button.classList.toggle('is-active',active);button.setAttribute('aria-pressed',active?'true':'false');}});
    }});
    updateCriterionLab();
}});
var criterionApply=document.getElementById('criterionApply');
if(criterionApply)criterionApply.addEventListener('click',criterionLabApplyResults);
updateCriterionLab();

// Combined outcome + tool filters. Multiple tools use AND semantics.
var activeStatusFilter='all';
var selectedToolFilters=new Set();

function applyCandidateFilters(){{
    var visibleCount=0;
    document.querySelectorAll('#tableBody .data-row').forEach(function(row){{
        var rowTools=new Set(String(row.dataset.tools||'').split(/\\s+/).filter(Boolean));
        var matchesStatus=activeStatusFilter==='all'||row.dataset.status===activeStatusFilter;
        var matchesTools=Array.from(selectedToolFilters).every(function(tool){{return rowTools.has(tool);}});
        var show=matchesStatus&&matchesTools;
        row.style.display=show?'':'none';
        row.classList.remove('selected');
        row.setAttribute('aria-expanded','false');
        if(show)visibleCount+=1;
    }});
    document.querySelectorAll('.detail-row').forEach(function(row){{row.style.display='none';}});
    var summary=document.getElementById('toolFilterSummary');
    if(summary){{
        var selectedLabels=Array.from(document.querySelectorAll('.tool-filter-chip[aria-pressed="true"] span')).map(function(node){{return node.textContent;}});
        summary.textContent='Showing '+visibleCount+' candidate'+(visibleCount===1?'':'s')+(selectedLabels.length?' with all: '+selectedLabels.join(' + '):'');
    }}
    var clear=document.getElementById('toolFilterClear');
    if(clear)clear.disabled=selectedToolFilters.size===0;
}}

document.querySelectorAll('.stat-box[data-f]').forEach(box=>{{
    box.addEventListener('click',()=>{{
        const f=box.dataset.f;
        activeStatusFilter=f;
        document.querySelectorAll('.stat-box[data-f]').forEach(s=>{{s.classList.remove('stat-active');s.setAttribute('aria-pressed','false');}});
        box.classList.add('stat-active');
        box.setAttribute('aria-pressed','true');
        applyCandidateFilters();
    }});
}});

document.querySelectorAll('.tool-filter-chip[data-tool]').forEach(function(chip){{
    chip.addEventListener('click',function(){{
        var tool=chip.dataset.tool;
        var selected=chip.getAttribute('aria-pressed')==='true';
        chip.setAttribute('aria-pressed',selected?'false':'true');
        if(selected)selectedToolFilters.delete(tool);else selectedToolFilters.add(tool);
        applyCandidateFilters();
    }});
}});

var toolFilterClear=document.getElementById('toolFilterClear');
if(toolFilterClear){{
    toolFilterClear.addEventListener('click',function(){{
        selectedToolFilters.clear();
        document.querySelectorAll('.tool-filter-chip').forEach(function(chip){{chip.setAttribute('aria-pressed','false');}});
        applyCandidateFilters();
    }});
}}

// Row click -> inline candidate review
function toggleCandidateReview(row){{
    var detail=row.nextElementSibling;
    var wasOpen=detail.style.display==='table-row';
    document.querySelectorAll('.detail-row').forEach(item=>item.style.display='none');
    document.querySelectorAll('.data-row').forEach(item=>{{
        item.classList.remove('selected');
        item.setAttribute('aria-expanded','false');
    }});
    if(!wasOpen){{
        detail.style.display='table-row';
        row.classList.add('selected');
        row.setAttribute('aria-expanded','true');
    }}
}}

document.querySelectorAll('.data-row').forEach(row=>{{
    row.addEventListener('click',()=>{{
        toggleCandidateReview(row);
    }});
    row.addEventListener('keydown',event=>{{
        if(event.key==='Enter'||event.key===' '){{
            event.preventDefault();
            toggleCandidateReview(row);
        }}
    }});
}});
document.querySelectorAll('.data-row a').forEach(link=>{{
    link.addEventListener('click',event=>event.stopPropagation());
}});

// Ambiguity-only AI verdict
var autoAiReview={"true" if auto_ai_review else "false"};
var aiProvider={json.dumps(ai_provider)};
var aiProviderName={json.dumps(ai_provider_name)};
var showAiUsage={"true" if show_ai_usage else "false"};
var hashToken=(location.hash.match(/token=([^&]+)/)||[])[1]||'';
var aiToken=hashToken?decodeURIComponent(hashToken):'';
try{{
    if(hashToken){{
        sessionStorage.setItem('kriterion-token',aiToken);
        history.replaceState(null,'',location.pathname+location.search);
    }}else{{
        aiToken=sessionStorage.getItem('kriterion-token')||'';
    }}
}}catch(e){{}}
function aiHeaders(){{return{{'Content-Type':'application/json','X-Kriterion-Token':aiToken}};}}

function showAiError(node,message){{
    node.innerHTML='';
    var error=document.createElement('div');
    error.className='ai-error';
    error.appendChild(semanticElement('strong','ai-error-title','AI verdict unavailable'));
    error.appendChild(semanticElement('span','ai-error-detail',message||(aiProviderName+' could not produce a usable response.')));
    node.appendChild(error);
}}

function semanticElement(tag,className,text){{
    var node=document.createElement(tag);
    node.className=className;
    node.textContent=String(text||'');
    return node;
}}

function showInterviewError(node,message){{
    node.innerHTML='';
    var error=semanticElement('div','ai-error','');
    error.appendChild(semanticElement('strong','ai-error-title','Interview plan unavailable'));
    error.appendChild(semanticElement('span','ai-error-detail',message||(aiProviderName+' could not produce a verified interview plan.')));
    node.appendChild(error);
}}

var interviewCategories=[
    {{key:'AMBIGUOUS_EXPERIENCE',title:'Ambiguous evidence',copy:'Uncertain scope, ownership, dates, or requirement evidence'}},
    {{key:'STRONG_CLAIM',title:'Strong claims',copy:'Quantified results that need their baseline, measurement, testing, and personal attribution verified'}},
    {{key:'CAREER_TIMELINE',title:'Timeline signals',copy:'Verified career gaps, overlaps, date conflicts, or unclear transitions'}}
];

function renderInterviewPlan(container,plan){{
    var output=container.querySelector('.interview-architect-output');
    output.innerHTML='';
    output.appendChild(semanticElement('div','interview-plan-summary',plan.summary));
    var groups=semanticElement('div','interview-groups','');
    var questions=Array.isArray(plan.questions)?plan.questions:[];
    var issueNumber=0;
    interviewCategories.forEach(function(category){{
        var items=questions.filter(function(item){{return item.category===category.key;}});
        var group=semanticElement('section','interview-group','');
        var head=semanticElement('div','interview-group-head','');
        var titleWrap=semanticElement('div','','');
        titleWrap.appendChild(semanticElement('div','interview-group-title',category.title));
        titleWrap.appendChild(semanticElement('span','interview-group-copy',category.copy));
        head.appendChild(titleWrap);
        head.appendChild(semanticElement('span','interview-group-count',String(items.length)));
        group.appendChild(head);
        if(!items.length){{
            group.appendChild(semanticElement('div','interview-no-signal','No defensible CV signal found — no question generated.'));
        }}else{{
            var list=semanticElement('div','interview-question-list','');
            items.forEach(function(item){{
                issueNumber+=1;
                var card=semanticElement('article','interview-question','');
                card.dataset.priority=item.priority;
                var top=semanticElement('div','interview-question-top','');
                top.appendChild(semanticElement('span','interview-question-number','ISSUE '+String(issueNumber).padStart(2,'0')));
                top.appendChild(semanticElement('span','interview-priority',item.priority+' PRIORITY'));
                card.appendChild(top);
                card.appendChild(semanticElement('span','interview-source-label','Detected issue'));
                card.appendChild(semanticElement('div','interview-question-text',item.issue||item.rationale));
                card.appendChild(semanticElement('div','interview-rationale',item.rationale));
                card.appendChild(semanticElement('span','interview-source-label','Single interview question'));
                card.appendChild(semanticElement('div','interview-question-text',item.question));
                var listen=semanticElement('div','interview-listen','');
                listen.appendChild(semanticElement('strong','','Listen for: '));
                listen.appendChild(document.createTextNode(String(item.what_to_listen_for||'')));
                card.appendChild(listen);
                card.appendChild(semanticElement('span','interview-source-label','CV evidence anchor'));
                (item.source_quotes||[]).forEach(function(quote){{
                    card.appendChild(semanticElement('blockquote','interview-source',quote));
                }});
                list.appendChild(card);
            }});
            group.appendChild(list);
        }}
        groups.appendChild(group);
    }});
    output.appendChild(groups);
    var usage=plan.token_usage;
    if(showAiUsage&&usage&&usage.available===true&&Number.isFinite(Number(usage.ai_credits))){{
        var calls=tokenCount(usage.ai_calls);
        var sharedPrefix=usage.shared_analysis===true?'Shared analysis · ':'';
        output.appendChild(semanticElement('div','interview-usage',sharedPrefix+formatCreditCount(usage.ai_credits)+' GitHub AI credit'+(Number(usage.ai_credits)===1?'':'s')+' · '+formatTokenCount(calls)+' AI call'+(calls===1?'':'s')));
    }}
}}

function requestInterviewPlan(btn){{return requestPassedCandidateAnalysis(btn);}}

var passedCandidateAnalysisState={{}};

function requestPassedCandidateAnalysis(btn){{
    var source=btn.closest('.interview-architect-container');
    var idx=source.dataset.idx;
    var file=source.dataset.file;
    var output=source.querySelector('.interview-architect-output');
    if(!aiToken){{
        showInterviewError(output,'Interview Architect requires the locally served dashboard. Run Kriterion again to open it.');
        return Promise.resolve();
    }}
    var state=passedCandidateAnalysisState[idx];
    if(!state){{
        state={{promise:null,data:null}};
        passedCandidateAnalysisState[idx]=state;
    }}

    if(state.data){{
        renderInterviewPlan(source,state.data.plan);
        btn.style.display='none';
        return Promise.resolve(state.data);
    }}
    btn.disabled=true;
    btn.textContent='Analyzing candidate...';
    output.innerHTML='<div class="ai-loading"><div class="spinner"></div><span>Detecting issues and building one question for each...</span></div>';
    if(state.promise)return state.promise;

    state.promise=fetch('/api/passed-candidate-analysis',{{method:'POST',headers:aiHeaders(),body:JSON.stringify({{filename:file}})}})
    .then(function(response){{return response.json().then(function(data){{if(!response.ok)throw new Error(data.error||'Failed');return data;}});}})
    .then(function(data){{
        state.data=data;
        state.promise=null;
        renderInterviewPlan(source,data.plan);
        btn.style.display='none';
        return data;
    }})
    .catch(function(error){{
        state.promise=null;
        showInterviewError(output,error.message);
        btn.disabled=false;
        btn.textContent='Try again';
    }});
    return state.promise;
}}

function updateHumanDecision(container,decision){{
    var status=container.querySelector('.semantic-decision-status');
    var dataRow=document.querySelector('.data-row[data-idx="'+container.dataset.idx+'"]');
    var pill=dataRow?dataRow.querySelector('.pill'):null;
    if(!status)return;
    if(decision==='PASS'||decision==='FAIL'){{
        status.textContent='Final human decision: '+decision;
        status.style.color=decision==='PASS'?'var(--green)':'var(--red)';
        if(pill){{
            pill.textContent='FINAL '+decision;
            pill.className='pill pill-'+decision.toLowerCase();
        }}
    }}else{{
        status.textContent='Human final decision required';
        status.style.color='var(--text2)';
    }}
}}

function tokenCount(value){{
    var parsed=Number(value);
    return Number.isFinite(parsed)&&parsed>=0?Math.round(parsed):0;
}}

function formatTokenCount(value){{
    return new Intl.NumberFormat().format(tokenCount(value));
}}

function creditCount(value){{
    var parsed=Number(value);
    return Number.isFinite(parsed)&&parsed>=0?parsed:0;
}}

function formatCreditCount(value){{
    return new Intl.NumberFormat(undefined,{{maximumFractionDigits:2}}).format(creditCount(value));
}}

function formatModelName(value){{
    var words=String(value||'').trim().replace(/[_-]+/g,' ').split(/\\s+/).filter(Boolean);
    return words.map(function(word){{
        var lower=word.toLowerCase();
        if(lower==='gpt')return 'GPT';
        if(lower==='claude')return 'Claude';
        if(lower==='opus')return 'Opus';
        if(lower==='sonnet')return 'Sonnet';
        if(lower==='haiku')return 'Haiku';
        if(lower==='gemini')return 'Gemini';
        if(/^\\d+(?:\\.\\d+)*$/.test(word))return word;
        return word.charAt(0).toUpperCase()+word.slice(1);
    }}).join(' ');
}}

function updateAmbiguityCreditTotal(){{
    var containers=Array.from(document.querySelectorAll('.semantic-review-container'));
    var total=0,loaded=0,unavailable=0,calls=0,models=new Set();
    containers.forEach(function(container){{
        if(container.dataset.creditUsageState==='available'){{
            loaded+=1;
            total+=creditCount(container.dataset.aiCredits);
            calls+=tokenCount(container.dataset.aiCalls);
            String(container.dataset.aiModels||'').split('|').filter(Boolean).forEach(function(model){{models.add(model);}});
        }}else if(container.dataset.creditUsageState==='unavailable'){{
            unavailable+=1;
        }}
    }});
    var totalNode=document.getElementById('aiUsageTotal');
    var detailNode=document.getElementById('aiUsageBreakdown');
    if(!totalNode||!detailNode)return;
    if(!loaded&&unavailable){{
        totalNode.textContent='Credit usage unavailable';
        detailNode.textContent=formatTokenCount(unavailable)+' cached review'+(unavailable===1?' has':'s have')+' no Copilot credit metadata.';
        return;
    }}
    if(!loaded){{
        totalNode.textContent='No AI reviews loaded';
        detailNode.textContent='Exact Copilot credit usage appears after ambiguous CV verdicts are generated or loaded.';
        return;
    }}
    totalNode.textContent=formatCreditCount(total)+' '+(total===1?'credit':'credits');
    var modelText=Array.from(models).map(formatModelName).join(', ');
    detailNode.textContent=(modelText?modelText+' · ':'')+formatTokenCount(loaded)+' CV review'+(loaded===1?'':'s')+' · '+formatTokenCount(calls)+' AI call'+(calls===1?'':'s')+(unavailable?' · '+formatTokenCount(unavailable)+' older review'+(unavailable===1?'':'s')+' unavailable':'')+'.';
}}

function renderCreditUsage(container,review,output){{
    var usage=review&&review.token_usage;
    var card=semanticElement('div','ai-token-usage','');
    var copy=semanticElement('div','','');
    var credits=usage?Number(usage.ai_credits):NaN;
    if(usage&&usage.available===true&&Number.isFinite(credits)&&credits>=0){{
        var calls=tokenCount(usage.ai_calls);
        var attempts=tokenCount(usage.attempts);
        var models=Array.isArray(usage.models)?usage.models.map(String).filter(Boolean):[];
        var modelLabel=models.length?models.map(formatModelName).join(', '):'Copilot';
        var details=[formatTokenCount(calls)+' AI call'+(calls===1?'':'s')];
        if(attempts>1)details.push(formatTokenCount(attempts)+' attempts');
        copy.appendChild(semanticElement('strong','',modelLabel+' • '+formatCreditCount(credits)+' '+(credits===1?'credit':'credits')));
        copy.appendChild(semanticElement('span','',details.join(' · ')));
        container.dataset.creditUsageState='available';
        container.dataset.aiCredits=String(credits);
        container.dataset.aiModels=models.join('|');
        container.dataset.aiCalls=String(calls);
    }}else{{
        card.className+=' ai-token-unavailable';
        copy.appendChild(semanticElement('strong','','Credit usage unavailable'));
        copy.appendChild(semanticElement('span','','Copilot did not return exact AI credit metadata for this review.'));
        container.dataset.creditUsageState='unavailable';
    }}
    card.appendChild(copy);
    card.appendChild(semanticElement('div','ai-token-usage-label','GitHub AI Credits'));
    output.appendChild(card);
    updateAmbiguityCreditTotal();
}}

function renderAmbiguityVerdict(container,review){{
    var output=container.querySelector('.semantic-review-output');
    output.innerHTML='';
    output.appendChild(semanticElement(
        'div',
        'ai-verdict '+String(review.ai_verdict||'').toLowerCase(),
        'AI verdict: '+review.ai_verdict
    ));
    output.appendChild(semanticElement('div','semantic-summary',review.summary));
    if(showAiUsage)renderCreditUsage(container,review,output);
    var recommendationReasons=Array.isArray(review.reasons)&&review.reasons.length?review.reasons:(review.evidence||[]).map(function(item){{
        return {{criterion:item.criterion,status:item.stance==='SUPPORTS_FAIL'?'FAILED':'MET',explanation:item.explanation,source_quote:item.source_quote,confidence:item.confidence}};
    }});
    recommendationReasons.forEach(function(finding){{
        var card=semanticElement('div','semantic-finding','');
        var head=semanticElement('div','semantic-finding-head','');
        head.appendChild(semanticElement('div','semantic-finding-title',finding.criterion));
        head.appendChild(semanticElement('span','semantic-classification',String(finding.status||'REASON').replaceAll('_',' ')));
        card.appendChild(head);
        card.appendChild(semanticElement('div','semantic-explanation',finding.explanation));
        if(finding.source_quote){{
            card.appendChild(semanticElement('blockquote','semantic-quote',finding.source_quote));
            var confidence=Number.isFinite(Number(finding.confidence))?'AI confidence: '+finding.confidence+'% · ':'';
            card.appendChild(semanticElement('div','semantic-meta',confidence+'quotation verified against parsed work experience'));
        }}else{{
            card.appendChild(semanticElement('div','semantic-meta','No citation applicable — this reason is based on an absent or numeric screening requirement.'));
        }}
        output.appendChild(card);
    }});
    var actions=semanticElement('div','semantic-actions','');
    var pass=semanticElement('button','semantic-decision pass','Final decision: Pass candidate');
    pass.type='button';
    pass.onclick=function(){{recordFinalDecision(pass,'PASS');}};
    var fail=semanticElement('button','semantic-decision fail','Final decision: Fail candidate');
    fail.type='button';
    fail.onclick=function(){{recordFinalDecision(fail,'FAIL');}};
    actions.appendChild(pass);
    actions.appendChild(fail);
    actions.appendChild(semanticElement('span','semantic-decision-status',''));
    output.appendChild(actions);
    updateHumanDecision(container,review.human_decision);
}}

function requestAmbiguityVerdict(btn){{
    var container=btn.closest('.semantic-review-container');
    var file=container.dataset.file;
    var output=container.querySelector('.semantic-review-output');
    if(!aiToken){{
        showAiError(output,'AI Verdict requires the locally served dashboard. Run Kriterion again to open it.');
        return Promise.resolve();
    }}
    btn.disabled=true;
    btn.textContent='Checking evidence...';
    output.innerHTML='<div class="ai-loading"><div class="spinner"></div><span>Checking the unresolved evidence...</span></div>';
    return fetch('/api/ai-verdict',{{method:'POST',headers:aiHeaders(),body:JSON.stringify({{filename:file}})}})
    .then(function(r){{return r.json().then(function(d){{if(!r.ok)throw new Error(d.error||'Failed');return d;}});}})
    .then(function(data){{
        renderAmbiguityVerdict(container,data.review);
        btn.style.display='none';
    }})
    .catch(function(e){{
        showAiError(output,e.message);
        btn.disabled=false;
        btn.textContent='Try again';
    }});
}}

function recordFinalDecision(btn,decision){{
    var container=btn.closest('.semantic-review-container');
    var file=container.dataset.file;
    var buttons=container.querySelectorAll('.semantic-decision');
    var status=container.querySelector('.semantic-decision-status');
    buttons.forEach(function(item){{item.disabled=true;}});
    status.textContent='Saving decision...';
    fetch('/api/final-decision',{{method:'POST',headers:aiHeaders(),body:JSON.stringify({{
        filename:file,decision:decision
    }})}})
    .then(function(r){{return r.json().then(function(d){{if(!r.ok)throw new Error(d.error||'Failed');return d;}});}})
    .then(function(data){{
        updateHumanDecision(container,data.human_decision);
        buttons.forEach(function(item){{item.disabled=false;}});
    }})
    .catch(function(e){{
        status.textContent=e.message;
        buttons.forEach(function(item){{item.disabled=false;}});
    }});
}}

function autoReviewAmbiguous(){{
    if(!autoAiReview||!aiToken)return;
    var buttons=Array.from(document.querySelectorAll('.semantic-review-btn'));
    var sequence=Promise.resolve();
    buttons.forEach(function(btn){{
        sequence=sequence.then(function(){{return requestAmbiguityVerdict(btn);}});
    }});
}}

function sendHeartbeat(){{
    if(aiToken)fetch('/heartbeat',{{headers:aiHeaders()}}).catch(function(){{}});
}}
sendHeartbeat();
window.setInterval(sendHeartbeat,10000);
window.setTimeout(autoReviewAmbiguous,200);
</script>
</body>
</html>"""

    output_path.write_text(html, encoding="utf-8")


def write_excel(results: List[Dict[str, object]], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Screening Results"

    headers = build_dynamic_headers()
    ws.append(headers)

    for r in results:
        ws.append(row_for_result(r, headers))

    header_font = Font(bold=True)
    top = Alignment(vertical="top")
    wrap_top = Alignment(wrap_text=True, vertical="top")

    ws.row_dimensions[1].height = 22
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = top

    ws.freeze_panes = "A2"

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    header_to_col = {headers[i]: i + 1 for i in range(len(headers))}

    snippet_cols = [h for h in headers if h.endswith("_snippet")]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = top
        for h in snippet_cols:
            col_idx = header_to_col[h]
            row[col_idx - 1].alignment = wrap_top

    result_col = header_to_col["result"]
    fill_pass = PatternFill(fill_type="solid", fgColor="00A000")
    fill_fail = PatternFill(fill_type="solid", fgColor="C00000")
    fill_amb = PatternFill(fill_type="solid", fgColor="F39C12")
    font_white = Font(color="FFFFFF", bold=True)

    for rr in range(2, ws.max_row + 1):
        cell = ws.cell(row=rr, column=result_col)
        val = (cell.value or "").strip().upper()
        if val == "PASS":
            cell.fill = fill_pass
            cell.font = font_white
        elif val == "FAIL":
            cell.fill = fill_fail
            cell.font = font_white
        elif val == "AMBIGUOUS":
            cell.fill = fill_amb
            cell.font = font_white
        cell.alignment = Alignment(horizontal="center", vertical="top")

    caps = {h: 45 for h in headers}
    caps["file"] = 55
    caps["result"] = 14
    caps["devops_years"] = 14
    for h in snippet_cols:
        caps[h] = 80

    for col_idx, h in enumerate(headers, start=1):
        longest = len(h)
        for rr in range(2, ws.max_row + 1):
            v = ws.cell(row=rr, column=col_idx).value
            if v is None:
                continue
            longest = max(longest, len(str(v)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            caps.get(h, 45), max(10, longest + 2)
        )

    snippet_col_indices = [header_to_col[h] for h in snippet_cols]
    for rr in range(2, ws.max_row + 1):
        long_snip = False
        for col in snippet_col_indices:
            val = ws.cell(row=rr, column=col).value or ""
            if len(str(val)) > 80:
                long_snip = True
                break
        ws.row_dimensions[rr].height = 70 if long_snip else 20

    wb.save(output_path)
